import streamlit as st
from supabase import create_client, Client
import base64
import calendar
import fcntl
import hashlib
import io
import json
import logging
import os
import random
import smtplib
import zipfile
from contextlib import contextmanager
from email.message import EmailMessage
from datetime import date, datetime, time, timedelta
from time import monotonic as _tiempo_monotonico, sleep as _dormir
from zoneinfo import ZoneInfo

import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from streamlit_js_eval import get_geolocation, streamlit_js_eval

# --- ZONA HORARIA (evita el desfase de horas del servidor, que corre en UTC) ---
ZONA_HORARIA_APP = ZoneInfo("America/Lima")


def ahora_peru() -> datetime:
    """Hora actual en la zona horaria de Perú (UTC-5), sin importar en qué
    huso horario esté el servidor donde corre la app."""
    return datetime.now(ZONA_HORARIA_APP)


def hoy_peru() -> date:
    """Fecha actual en la zona horaria de Perú."""
    return ahora_peru().date()


# Contraseña por defecto para empleados nuevos/de ejemplo. Configurable por
# despliegue vía el secret PASSWORD_EMPLEADO_DEFAULT; si no se configura,
# usa "123456" como respaldo.
PASSWORD_EMPLEADO_DEFAULT = st.secrets.get("PASSWORD_EMPLEADO_DEFAULT", "123456")

# --- CONEXIÓN SEGURA A SUPABASE (NUBE EFÍMERA) ---
@st.cache_resource
def init_supabase() -> Client:
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        st.warning("⚠️ No se pudo conectar a la Nube (Supabase). Trabajando en modo 100% Local.")
        return None

supabase = init_supabase()

# Función para enviar marcaciones a Supabase (sincronización remota)
def ya_marco_hoy(supabase, empresa_id, dni, nombre, fecha_str, tipo_marcacion):
    """Devuelve True si el empleado ya tiene esa marcación (Entrada/Salida)
    registrada hoy. Consulta Supabase primero (fuente compartida entre
    dispositivos); si no está disponible, revisa el CSV local como respaldo."""
    if supabase:
        try:
            res = (
                supabase.table("marcaciones_efimeras")
                .select("id")
                .eq("empresa_id", str(empresa_id))
                .eq("dni", str(dni))
                .eq("fecha", str(fecha_str))
                .eq("tipo", str(tipo_marcacion))
                .limit(1)
                .execute()
            )
            return len(res.data) > 0
        except Exception:
            pass  # si falla la consulta a la nube, se revisa el CSV local

    if os.path.exists(CSV_ASISTENCIA):
        try:
            df_check = pd.read_csv(CSV_ASISTENCIA)
            coincidencias = df_check[
                (df_check["empresa_id"].astype(str) == str(empresa_id))
                & (df_check["Empleado"] == nombre)
                & (df_check["Fecha"].astype(str) == str(fecha_str))
                & (df_check["Tipo Marcación"] == tipo_marcacion)
            ]
            return not coincidencias.empty
        except Exception:
            return False

    return False


def validar_foto_captura(img_file, max_mb=8):
    """Valida la foto tomada con la cámara antes de guardarla y subirla:
    que exista, que no pese más de la cuenta, y que sea una imagen
    válida y no esté dañada. Devuelve (es_valida, mensaje_de_error)."""
    if img_file is None:
        return False, "No se detectó ninguna foto."
    if img_file.size == 0:
        return False, "La foto capturada está vacía. Vuelve a intentarlo."

    tam_mb = img_file.size / (1024 * 1024)
    if tam_mb > max_mb:
        return False, (
            f"La foto pesa {tam_mb:.1f} MB, más del máximo permitido"
            f" ({max_mb} MB). Vuelve a tomarla."
        )

    try:
        img_file.seek(0)
        imagen_prueba = Image.open(img_file)
        imagen_prueba.verify()
    except Exception:
        return False, (
            "El archivo de la foto está dañado o no es una imagen"
            " válida. Vuelve a tomarla."
        )
    finally:
        img_file.seek(0)

    return True, ""


def enviar_marcacion_supabase(empresa_id, dni, nombre, fecha, hora, tipo, foto_url="", gps=""):
    if not supabase:
        return False
    try:
        data = {
            "empresa_id": str(empresa_id),
            "empleado_dni": str(dni),
            "empleado_nombre": str(nombre),
            "fecha": str(fecha),
            "hora": str(hora),
            "tipo_marcacion": str(tipo),
            "foto_url": str(foto_url),
            "ubicacion_gps": str(gps)
        }
        supabase.table("marcaciones_efimeras").insert(data).execute()
        return True
    except Exception as e:
        print(f"Error al sincronizar con Supabase: {e}")
        return False

# ---------------------------------------------------------
# 1. Configuración de página e inicialización
# ---------------------------------------------------------

# Logger: los errores que antes se silenciaban del todo (except/pass)
# ahora quedan registrados aquí. No se le muestra nada al usuario (la
# experiencia no cambia), pero tú puedes revisarlos en Streamlit Cloud →
# tu app → menú (⋮) → "Manage app" → pestaña de Logs, para diagnosticar
# fallas que hoy pasan totalmente desapercibidas.
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("asistencia_app")


# --- Hash de PINs y contraseñas ---
# Los PINs (Admin/Visor/Master) y las contraseñas de trabajadores ya NO
# se guardan como texto plano: se guarda su hash (SHA-256 + sal fija de
# la app). Así, si alguna vez alguien accede directo a la base de datos
# de Supabase, no puede leer los PINs/contraseñas reales, solo el hash
# (que no se puede "revertir" a la clave original).
# La CLAVE DE EXCEL es la única excepción a propósito: openpyxl necesita
# el texto real para proteger la hoja de cálculo, así que esa se sigue
# guardando tal cual (igual nunca sale del entorno del developer).
_SAL_HASH = "asistencia_saas_v1"


def _hash_clave(valor_plano):
    """Convierte un PIN/contraseña en texto plano a su hash SHA-256."""
    return hashlib.sha256(
        f"{_SAL_HASH}:{valor_plano}".encode("utf-8")
    ).hexdigest()


def clave_coincide(valor_ingresado, valor_guardado):
    """Compara lo que la persona escribió contra lo guardado. Compatible
    con instalaciones viejas que todavía tengan el valor en texto plano
    (factory PINs, o Secrets de Streamlit): si lo guardado no tiene forma
    de hash SHA-256 (64 caracteres hexadecimales), se compara como texto
    plano; si sí la tiene, se compara el hash. Esto permite migrar sin
    romper logins existentes — en cuanto alguien guarda una clave nueva
    desde el panel, esa clave queda hasheada para siempre."""
    if valor_guardado is None:
        return False
    valor_guardado = str(valor_guardado)
    parece_hash = len(valor_guardado) == 64 and all(
        c in "0123456789abcdef" for c in valor_guardado.lower()
    )
    if parece_hash:
        return _hash_clave(valor_ingresado) == valor_guardado
    return str(valor_ingresado) == valor_guardado


st.set_page_config(
    page_title="Sistema de Asistencia y Nómina SaaS Multi-Empresa",
    layout="wide",
    page_icon="⏰",
)

st.markdown(
    """
    <link rel="icon" href="/app/static/icon-192.png">
    <link rel="apple-touch-icon" href="/app/static/icon-192.png">
    """,
    unsafe_allow_html=True,
)

# --- MODO MÓVIL / TRABAJADOR (vista simplificada, solo marcación) ---
# Se activa agregando ?modo=movil a la URL (opcionalmente también
# ?empresa=CODIGO para pre-cargar la empresa) — usado por la PWA instalada
# — O AUTOMÁTICAMENTE si se detecta que la pantalla es de tamaño celular.
MODO_MOVIL = st.query_params.get("modo") == "movil"
EMPRESA_URL = st.query_params.get("empresa")

if "ancho_pantalla_px" not in st.session_state:
    st.session_state.ancho_pantalla_px = None

_ancho_detectado = streamlit_js_eval(
    js_expressions="window.innerWidth", key="ANCHO_PANTALLA_PX"
)
if _ancho_detectado is not None:
    st.session_state.ancho_pantalla_px = _ancho_detectado

# ES_CELULAR: el dispositivo físico es un celular (por ancho de pantalla o
# por venir de la PWA), sin importar el rol de quien lo usa.
ES_CELULAR = MODO_MOVIL or (
    st.session_state.ancho_pantalla_px is not None
    and st.session_state.ancho_pantalla_px < 768
)

# VISTA_TRABAJADOR_MOVIL: además de ser celular, la persona todavía no
# inició sesión como Admin/SuperAdmin/Developer con PIN. Es la vista
# simplificada de solo marcación (sin menú lateral ni panel admin).
VISTA_TRABAJADOR_MOVIL = ES_CELULAR and not st.session_state.get(
    "autenticado", False
)

if VISTA_TRABAJADOR_MOVIL:
    st.markdown(
        """
        <link rel="manifest" href="/app/static/manifest.json">
        <meta name="theme-color" content="#111319">
        <meta name="apple-mobile-web-app-capable" content="yes">
        <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
        <link rel="apple-touch-icon" href="/app/static/icon-192.png">
        <style>
            [data-testid="stSidebar"] {display: none;}
            [data-testid="stToolbar"] {display: none;}
            footer {display: none;}
            #MainMenu {display: none;}
        </style>
        <script>
        if ('serviceWorker' in navigator) {
            navigator.serviceWorker.register('/app/static/service-worker.js');
        }
        </script>
        """,
        unsafe_allow_html=True,
    )

# ---------------------------------------------------------
# TEMA VISUAL GLOBAL — "futurista" (Parte 1 del rediseño)
# ---------------------------------------------------------
# Solo CSS/HTML: no cambia ningún widget, dato ni lógica de negocio.
# Reutiliza los mismos st.button/st.text_input/st.selectbox/etc. de
# siempre, solo les cambia la piel. El fondo animado va en una capa fija
# detrás de todo (z-index -1) para no interferir con los clics.
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Space+Mono:wght@400;700&display=swap');

    :root{
        --bg-base:#0a0c11;
        --bg-panel: rgba(19, 23, 32, 0.72);
        --bg-panel-solid:#12151d;
        --border: rgba(88, 166, 255, 0.18);
        --border-soft:#232733;
        --cyan:#58a6ff;
        --violet:#a371f7;
        --success:#3fb950;
        --amber:#d29922;
        --danger:#f85149;
        --text:#e6edf3;
        --text-muted:#8b949e;
        --font-display:'Space Grotesk', system-ui, -apple-system, sans-serif;
        --font-mono:'Space Mono', ui-monospace, SFMono-Regular, monospace;
    }

    /* ---------- FONDO FUTURISTA (capa fija detrás de todo) ---------- */
    .fac-bg-layer{position:fixed; inset:0; z-index:-1; overflow:hidden; pointer-events:none;}
    .fac-bg-base{
        position:absolute; inset:0;
        background:
            radial-gradient(ellipse 120% 80% at 50% -10%, rgba(88,166,255,0.10), transparent 60%),
            linear-gradient(180deg, #0a0c11 0%, #0d1017 55%, #0a0c11 100%);
    }
    .fac-orb{position:absolute; border-radius:50%; filter:blur(60px); opacity:0.35; will-change:transform;}
    .fac-orb-cyan{width:420px; height:420px; left:-10%; top:8%; background:radial-gradient(circle, var(--cyan), transparent 70%); animation:fac-drift1 22s ease-in-out infinite alternate;}
    .fac-orb-violet{width:380px; height:380px; right:-8%; top:35%; background:radial-gradient(circle, var(--violet), transparent 70%); animation:fac-drift2 26s ease-in-out infinite alternate;}
    .fac-orb-cyan2{width:300px; height:300px; left:20%; bottom:-10%; background:radial-gradient(circle, var(--cyan), transparent 70%); opacity:0.2; animation:fac-drift3 30s ease-in-out infinite alternate;}
    @keyframes fac-drift1{ from{transform:translate(0,0) scale(1);} to{transform:translate(60px,40px) scale(1.15);} }
    @keyframes fac-drift2{ from{transform:translate(0,0) scale(1);} to{transform:translate(-50px,-30px) scale(1.1);} }
    @keyframes fac-drift3{ from{transform:translate(0,0) scale(1);} to{transform:translate(40px,-50px) scale(1.2);} }
    .fac-grid-overlay{
        position:absolute; inset:-2px;
        background-image:
            linear-gradient(rgba(88,166,255,0.05) 1px, transparent 1px),
            linear-gradient(90deg, rgba(88,166,255,0.05) 1px, transparent 1px);
        background-size:42px 42px;
        mask-image:radial-gradient(ellipse 80% 60% at 50% 30%, black 40%, transparent 85%);
        opacity:0.5;
    }
    .fac-scanline{
        position:absolute; left:0; right:0; height:120px;
        background:linear-gradient(180deg, transparent, rgba(88,166,255,0.06), transparent);
        animation:fac-scan 7s linear infinite;
    }
    @keyframes fac-scan{ 0%{top:-120px;} 100%{top:100%;} }
    @media (prefers-reduced-motion: reduce){ .fac-orb, .fac-scanline{ animation:none !important; } }

    /* ---------- APLICAR EL TEMA A LOS WIDGETS NATIVOS DE STREAMLIT ---------- */
    [data-testid="stAppViewContainer"], [data-testid="stHeader"]{ background:transparent !important; }
    .stApp{ background:transparent !important; }
    body, [data-testid="stAppViewContainer"] { color:var(--text); }

    h1, h2, h3{ font-family:var(--font-display) !important; letter-spacing:0.01em; }
    h1{
        background:linear-gradient(90deg, #fff, var(--cyan) 120%);
        -webkit-background-clip:text; background-clip:text;
    }

    /* Botones */
    .stButton > button{
        font-family:var(--font-display);
        font-weight:600;
        border-radius:14px !important;
        border:1px solid var(--border-soft) !important;
        background:var(--bg-panel-solid) !important;
        color:var(--text) !important;
        transition:transform 0.15s ease, border-color 0.15s ease, box-shadow 0.15s ease;
    }
    .stButton > button:hover{
        border-color:var(--cyan) !important;
        transform:translateY(-1px);
    }
    /* Botón "primario" (type="primary"): el llamado a la acción grande,
       usado en Confirmar Marcación y otros botones clave del sistema. */
    .stButton > button[kind="primary"]{
        background:linear-gradient(135deg, #6fb4ff, var(--cyan) 45%, var(--violet) 130%) !important;
        border:none !important;
        color:#fff !important;
        font-size:1.05em;
        padding:0.7em 1.4em !important;
        box-shadow:0 12px 34px rgba(88,166,255,0.35), 0 4px 14px rgba(0,0,0,0.4);
    }
    .stButton > button[kind="primary"]:hover{
        transform:translateY(-2px);
        box-shadow:0 16px 40px rgba(88,166,255,0.45), 0 4px 14px rgba(0,0,0,0.4);
    }
    .stButton > button[kind="primary"]:disabled{
        background:var(--bg-panel-solid) !important;
        color:var(--text-muted) !important;
        box-shadow:none;
    }

    /* Tarjetas / contenedores con borde (st.container(border=True)) */
    [data-testid="stVerticalBlockBorderWrapper"]{
        background:var(--bg-panel) !important;
        backdrop-filter:blur(18px);
        -webkit-backdrop-filter:blur(18px);
        border:1px solid var(--border) !important;
        border-radius:22px !important;
    }

    /* Inputs */
    [data-testid="stTextInput"] input, [data-testid="stNumberInput"] input,
    [data-baseweb="select"] > div, [data-baseweb="input"]{
        background:var(--bg-panel-solid) !important;
        border-radius:12px !important;
        border-color:var(--border-soft) !important;
        color:var(--text) !important;
    }

    /* Alertas: se conserva el color semántico de Streamlit, solo se
       redondea y se le da aire "glass" sin tocar el color de fondo. */
    [data-testid="stAlert"]{
        border-radius:14px !important;
        backdrop-filter:blur(6px);
    }

    /* Expanders */
    [data-testid="stExpander"]{
        border-radius:16px !important;
        border-color:var(--border-soft) !important;
        background:var(--bg-panel-solid) !important;
    }

    /* Pestañas (tabs) del panel admin */
    [data-baseweb="tab-list"]{ gap:4px; }
    [data-baseweb="tab"]{
        font-family:var(--font-display);
        border-radius:10px 10px 0 0 !important;
    }

    /* Sidebar */
    [data-testid="stSidebar"]{
        background:rgba(13,16,23,0.9) !important;
        border-right:1px solid var(--border-soft);
    }

    /* Métricas (st.metric) */
    [data-testid="stMetric"]{
        background:var(--bg-panel-solid);
        border:1px solid var(--border-soft);
        border-radius:14px;
        padding:10px 14px;
    }
    </style>

    <div class="fac-bg-layer">
        <div class="fac-bg-base"></div>
        <div class="fac-orb fac-orb-cyan"></div>
        <div class="fac-orb fac-orb-violet"></div>
        <div class="fac-orb fac-orb-cyan2"></div>
        <div class="fac-grid-overlay"></div>
        <div class="fac-scanline"></div>
    </div>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    """
<style>
    .bitacora-container {
        background-color: #111319;
        border-radius: 12px;
        padding: 16px;
        border: 1px solid #232733;
        box-shadow: 0 4px 20px rgba(0,0,0,0.5);
        margin-top: 15px;
        margin-bottom: 20px;
    }
    table.bitacora-table {
        width: 100%;
        border-collapse: collapse;
        color: #c9d1d9;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
        font-size: 13px;
    }
    table.bitacora-table th {
        background-color: #161b22;
        color: #8b949e;
        padding: 12px 10px;
        text-align: left;
        font-weight: 600;
        border-bottom: 2px solid #30363d;
    }
    table.bitacora-table td {
        padding: 10px;
        border-bottom: 1px solid #21262d;
        vertical-align: middle;
    }
    table.bitacora-table tr:hover {
        background-color: #161b22;
    }
    tr.week-divider td {
        border-top: 3px solid #1f6feb !important;
        background-color: #0d1117;
    }
    .user-avatar-thumb {
        width: 42px;
        height: 42px;
        object-fit: cover;
        border-radius: 8px;
        border: 1px solid #3a3f4d;
        display: block;
        cursor: pointer;
        transition: transform 0.15s ease, border-color 0.15s ease;
    }
    .user-avatar-thumb:hover {
        border-color: #58a6ff;
        transform: scale(1.05);
    }
    .img-modal-backdrop {
        position: fixed;
        z-index: 999999;
        left: 0;
        top: 0;
        width: 100vw;
        height: 100vh;
        background-color: rgba(0, 0, 0, 0.85);
        backdrop-filter: blur(5px);
        display: flex;
        align-items: center;
        justify-content: center;
        opacity: 0;
        pointer-events: none;
        transition: opacity 0.2s ease-in-out;
    }
    .img-modal-backdrop:target {
        opacity: 1;
        pointer-events: auto;
    }
    .img-modal-content {
        position: relative;
        max-width: 85vw;
        max-height: 85vh;
        display: flex;
        justify-content: center;
        align-items: center;
    }
    .img-modal-content img {
        max-width: 85vw;
        max-height: 85vh;
        border-radius: 12px;
        box-shadow: 0 0 30px rgba(0, 0, 0, 0.9);
        border: 1px solid #30363d;
    }
    .img-modal-close {
        position: absolute;
        top: -15px;
        right: -15px;
        width: 36px;
        height: 36px;
        background-color: #f85149;
        color: #ffffff !important;
        border-radius: 50%;
        text-decoration: none !important;
        font-weight: bold;
        font-size: 18px;
        line-height: 34px;
        text-align: center;
        box-shadow: 0 2px 10px rgba(0,0,0,0.5);
        transition: background-color 0.2s, transform 0.15s;
        border: 2px solid #ffffff;
        z-index: 1000000;
    }
    .img-modal-close:hover {
        background-color: #da3633;
        transform: scale(1.1);
    }
    .img-modal-close-overlay {
        position: absolute;
        width: 100%;
        height: 100%;
        top: 0;
        left: 0;
        z-index: -1;
    }
    .badge-puntual { color: #3fb950; background-color: rgba(46, 160, 67, 0.15); padding: 4px 8px; border-radius: 12px; font-weight: 600; display: inline-block; }
    .badge-tardanza { color: #f85149; background-color: rgba(248, 81, 73, 0.15); padding: 4px 8px; border-radius: 12px; font-weight: 600; display: inline-block; }
    .badge-falta { color: #ff4d4d; background-color: rgba(255, 77, 77, 0.15); padding: 4px 8px; border-radius: 12px; font-weight: 600; display: inline-block; }
    .badge-feriado { color: #58a6ff; background-color: rgba(88, 166, 255, 0.15); padding: 4px 8px; border-radius: 12px; font-weight: 600; display: inline-block; }
</style>

<script>
document.addEventListener('keydown', function(event) {
    if (event.key === "Escape") {
        if (window.location.hash.startsWith("#img-modal-")) {
            window.location.hash = "#_";
        }
    }
});
</script>
""",
    unsafe_allow_html=True,
)

if "entorno" not in st.session_state:
    st.session_state.entorno = "PROD"
if "empresa_id" not in st.session_state:
    st.session_state.empresa_id = "EMP_DEMO"
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "rol" not in st.session_state:
    st.session_state.rol = None
if "pin_admin" not in st.session_state:
    st.session_state.pin_admin = st.secrets.get("PIN_ADMIN", "1234")
if "pin_visor" not in st.session_state:
    st.session_state.pin_visor = st.secrets.get("PIN_VISOR", "5678")
if "pin_master" not in st.session_state:
    st.session_state.pin_master = st.secrets.get("PIN_MASTER", "9999")
if "clave_excel" not in st.session_state:
    st.session_state.clave_excel = st.secrets.get("CLAVE_EXCEL", "admin123")
if "fecha_inicio_sistema" not in st.session_state:
    st.session_state.fecha_inicio_sistema = date(2026, 1, 1)

# Estado persistente para la activación del despliegue en producción
if "mejoras_activadas_prod" not in st.session_state:
    st.session_state.mejoras_activadas_prod = False

if "emp_login_ok" not in st.session_state:
    st.session_state.emp_login_ok = False
if "emp_datos" not in st.session_state:
    st.session_state.emp_datos = None

if "dias_laborables" not in st.session_state:
    st.session_state.dias_laborables = [
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
    ]
if "tolerancia_minutos" not in st.session_state:
    st.session_state.tolerancia_minutos = 5
if "fecha_inicio_tolerancia" not in st.session_state:
    st.session_state.fecha_inicio_tolerancia = date(2026, 1, 1)

FERIADOS_OFICIALES = {
    "2026-01-01": "Año Nuevo",
    "2026-04-02": "Jueves Santo",
    "2026-04-03": "Viernes Santo",
    "2026-05-01": "Día del Trabajo",
    "2026-06-07": "Batalla de Arica",
    "2026-06-29": "San Pedro y San Pablo",
    "2026-07-23": "Día de la Fuerza Aérea",
    "2026-07-28": "Fiestas Patrias",
    "2026-07-29": "Fiestas Patrias",
    "2026-08-06": "Batalla de Junín",
    "2026-08-30": "Santa Rosa de Lima",
    "2026-10-08": "Combate de Angamos",
    "2026-11-01": "Día de Todos los Santos",
    "2026-12-08": "Inmaculada Concepción",
    "2026-12-09": "Batalla de Ayacucho",
    "2026-12-25": "Navidad",
}

# ---------------------------------------------------------
# 2. Persistencia de Datos y Archivos CSV
# ---------------------------------------------------------
CSV_EMPRESAS = "empresas.csv"
CSV_ASISTENCIA = "asistencia_multitienda.csv"
CSV_SEDES = "sedes.csv"
CSV_EMPLEADOS = "empleados.csv"
DIR_FOTOS = "fotos_asistencia"

if not os.path.exists(DIR_FOTOS):
    os.makedirs(DIR_FOTOS)


@contextmanager
def bloqueo_csv(ruta_csv, timeout=15):
    """Bloqueo exclusivo de archivo (a nivel de sistema operativo, vía
    fcntl) para evitar que dos operaciones simultáneas sobre el mismo CSV
    se pisen entre sí — por ejemplo, dos trabajadores marcando asistencia
    al mismo tiempo, o un admin editando mientras se sincroniza desde la
    nube.

    IMPORTANTE para quien edite este código: hay que usarlo envolviendo
    TODO el bloque que lee, modifica y vuelve a escribir el archivo — no
    solo la línea de escritura final. Si el bloqueo solo cubriera el
    to_csv(), dos hilos podrían leer el mismo estado viejo antes de que el
    otro termine de escribir, y el que escriba último borraría el cambio
    del primero (esto se llama "lost update" / carrera de datos).
    """
    ruta_lock = f"{ruta_csv}.lock"
    inicio = _tiempo_monotonico()
    f_lock = open(ruta_lock, "w")
    try:
        while True:
            try:
                fcntl.flock(f_lock, fcntl.LOCK_EX | fcntl.LOCK_NB)
                break
            except OSError:
                if _tiempo_monotonico() - inicio > timeout:
                    raise TimeoutError(
                        f"No se pudo obtener el bloqueo de {ruta_csv} a"
                        " tiempo (otra operación lo está usando)."
                    )
                _dormir(0.1)
        try:
            yield
        finally:
            fcntl.flock(f_lock, fcntl.LOCK_UN)
    finally:
        f_lock.close()

MESES_NOMBRES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
MESES_INVERSO = {v: k for k, v in MESES_NOMBRES.items()}
DIAS_SEMANA_MAP = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}


def es_mejora_activa(empresa_entorno):
    if empresa_entorno == "DEV":
        return True
    return st.session_state.get("mejoras_activadas_prod", False)


# Fila especial dentro de "configuracion_sistema" que no representa a
# ninguna empresa real: se usa solo para guardar el interruptor global
# de "Mejoras en Producción" (afecta a TODAS las empresas en PROD).
FLAG_GLOBAL_ID = "__SISTEMA_GLOBAL__"


def cargar_mejoras_prod_global(supabase):
    """Trae de Supabase si las 'Mejoras en Producción' están activas
    globalmente y actualiza session_state. Antes esto solo vivía en
    session_state: cada usuario/pestaña veía su propio estado y se
    perdía al cerrar el navegador o al redesplegar la app."""
    if not supabase:
        return
    try:
        res = (
            supabase.table("configuracion_sistema")
            .select("mejoras_activadas_prod")
            .eq("empresa_id", FLAG_GLOBAL_ID)
            .limit(1)
            .execute()
        )
        if res.data:
            valor = res.data[0].get("mejoras_activadas_prod")
            if valor is not None:
                st.session_state.mejoras_activadas_prod = bool(valor)
    except Exception:
        pass  # si falla, se sigue usando lo que ya había en session_state


def guardar_mejoras_prod_global(supabase, activo: bool):
    """Guarda en Supabase el estado global de 'Mejoras en Producción', para
    que sea el mismo para todas las empresas y dispositivos, y sobreviva
    a los redespliegues."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    supabase.table("configuracion_sistema").upsert(
        {"empresa_id": FLAG_GLOBAL_ID, "mejoras_activadas_prod": activo},
        on_conflict="empresa_id",
    ).execute()


def validar_integridad_desarrollo():
    if not os.path.exists(CSV_EMPRESAS):
        return False, "Falta archivo de empresas"
    
    df_emp = pd.read_csv(CSV_EMPRESAS)
    if "entorno" not in df_emp.columns or df_emp.empty:
        return False, "Estructura de empresas inválida"

    dev_count = len(df_emp[df_emp["entorno"] == "DEV"])
    if dev_count == 0:
        return False, "No existen cuentas de desarrollo registradas"

    return True, "Entorno de desarrollo completamente verificado"


@st.dialog("⚠️ Confirmación Requerida")
def modal_confirmar_despliegue():
    valido, mensaje = validar_integridad_desarrollo()
    
    if not valido:
        st.error(f"Error de validación: {mensaje}")
        st.write("Corrige la estructura en DEV antes de intentar la activación.")
        if st.button("Cerrar", use_container_width=True):
            st.rerun()
        return

    st.success(f"✓ {mensaje}")
    st.error("¡Atención! Esta acción desplegará los cambios a PRODUCCIÓN.")
    st.write(
        "Al activar, la nueva configuración de la suite de desarrollo se "
        "aplicará a **todas las empresas registradas en Producción**."
    )
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔴 Confirmar y Desplegar", use_container_width=True):
            if supabase:
                try:
                    guardar_mejoras_prod_global(supabase, True)
                except Exception as e:
                    st.warning(
                        "No se pudo guardar en la nube"
                        f" ({e}). Quedó activo solo para esta sesión;"
                        " otros usuarios/dispositivos no lo verán activo."
                    )
            st.session_state.mejoras_activadas_prod = True
            st.success("¡Despliegue a Producción exitoso!")
            st.rerun()
    with col2:
        if st.button("Cancelar", use_container_width=True):
            st.rerun()


def marcar_estado_modo_local(nombre_tabla: str, en_modo_local: bool):
    """Registra en session_state si una tabla (sedes/empleados/empresas)
    está operando en modo local (CSV) porque Supabase no respondió al
    CARGAR. Alimenta el aviso visible que se muestra en el panel admin
    (ver 'tablas_en_modo_local')."""
    tablas = st.session_state.setdefault("tablas_en_modo_local", set())
    if en_modo_local:
        tablas.add(nombre_tabla)
    else:
        tablas.discard(nombre_tabla)


def cargar_empresas_supabase(supabase):
    """Trae la lista completa de empresas desde Supabase. Devuelve None si
    Supabase no está disponible o falla la consulta (así quien llama sabe
    que debe usar el CSV local como respaldo/modo offline)."""
    if not supabase:
        return None
    try:
        res = supabase.table("empresas").select("*").execute()
        return res.data
    except Exception:
        return None


def guardar_empresa_supabase(supabase, datos_empresa):
    """Crea o actualiza (upsert) una empresa en Supabase."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    datos = dict(datos_empresa)
    datos["empresa_id"] = str(datos["empresa_id"])
    supabase.table("empresas").upsert(
        datos, on_conflict="empresa_id"
    ).execute()


def eliminar_empresa_supabase(supabase, empresa_id):
    """Borra (DELETE real) una empresa de Supabase."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    supabase.table("empresas").delete().eq(
        "empresa_id", str(empresa_id)
    ).execute()


def cargar_empresas():
    registros_empresas = cargar_empresas_supabase(supabase)
    columnas_empresas = [
        "empresa_id",
        "razon_social",
        "ruc",
        "plan",
        "estado",
        "entorno",
    ]

    if registros_empresas is not None:
        # Supabase respondió: es la fuente de verdad para la lista de
        # empresas (sobrevive a reboots/redespliegues, a diferencia del
        # CSV local que se borra en cada uno).
        if registros_empresas:
            df = pd.DataFrame(registros_empresas)
        else:
            df = pd.DataFrame(columns=columnas_empresas)

        valores_por_defecto = {
            "entorno": "PROD",
            "estado": "ACTIVO",
            "plan": "BASIC",
            "ruc": "",
            "razon_social": "",
        }
        for columna, valor_default in valores_por_defecto.items():
            if columna not in df.columns:
                df[columna] = valor_default
            df[columna] = df[columna].fillna(valor_default)

        # GARANTÍA: siempre debe existir al menos una empresa en entorno
        # DEV. Si no hubiera ninguna (tabla recién creada en Supabase, o
        # si se borraron todas por error), no habría forma de entrar al
        # panel Developer para dar de alta las demás empresas. Se crea
        # automáticamente y se guarda en Supabase para que quede fija.
        if df.empty or not (df["entorno"] == "DEV").any():
            empresa_dev_default = {
                "empresa_id": "DEV_TEST",
                "razon_social": "ENTORNO PRUEBAS DEV",
                "ruc": "20000000001",
                "plan": "DEVELOPER",
                "estado": "ACTIVO",
                "entorno": "DEV",
            }
            if supabase:
                try:
                    guardar_empresa_supabase(supabase, empresa_dev_default)
                except Exception as _e_silenciosa:
                    logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")
            df = pd.concat(
                [df, pd.DataFrame([empresa_dev_default])],
                ignore_index=True,
            )

        # Copia local como caché/respaldo por si Supabase falla más tarde.
        try:
            with bloqueo_csv(CSV_EMPRESAS):
                df[columnas_empresas].to_csv(CSV_EMPRESAS, index=False)
        except Exception as _e_silenciosa:
            logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")
        marcar_estado_modo_local("empresas", False)
        return df

    marcar_estado_modo_local("empresas", True)
    if os.path.exists(CSV_EMPRESAS):
        with bloqueo_csv(CSV_EMPRESAS):
            df = pd.read_csv(CSV_EMPRESAS)
            if "entorno" not in df.columns:
                df["entorno"] = "PROD"
                df.loc[df["empresa_id"] == "DEV_TEST", "entorno"] = "DEV"
            if df.empty or not (df["entorno"] == "DEV").any():
                df = pd.concat(
                    [
                        df,
                        pd.DataFrame([{
                            "empresa_id": "DEV_TEST",
                            "razon_social": "ENTORNO PRUEBAS DEV",
                            "ruc": "20000000001",
                            "plan": "DEVELOPER",
                            "estado": "ACTIVO",
                            "entorno": "DEV",
                        }]),
                    ],
                    ignore_index=True,
                )
            df.to_csv(CSV_EMPRESAS, index=False)
        return df
    else:
        df_init = pd.DataFrame({
            "empresa_id": ["EMP_DEMO", "DEV_TEST"],
            "razon_social": ["EMPRESA DEMO S.A.C.", "ENTORNO PRUEBAS DEV"],
            "ruc": ["20123456789", "20000000001"],
            "plan": ["PREMIUM", "DEVELOPER"],
            "estado": ["ACTIVO", "ACTIVO"],
            "entorno": ["PROD", "DEV"],
        })
        with bloqueo_csv(CSV_EMPRESAS):
            df_init.to_csv(CSV_EMPRESAS, index=False)
        return df_init


COLUMNAS_ASISTENCIA = [
    "empresa_id",
    "Fecha",
    "Empleado",
    "Tipo Marcación",
    "Hora Registrada",
    "Hora Entrada Oficial",
    "Hora Salida Oficial",
    "Estado",
    "Minutos Tardanza",
    "Horas Extra (min)",
    "Sede Detectada",
    "Distancia (m)",
    "En Rango",
    "Foto",
]


def sincronizar_marcaciones_nube(supabase, empresa_id):
    """Trae desde Supabase las marcaciones de esta empresa de los últimos
    días que aún no estén en el CSV local, y las agrega. Se llama cada vez
    que se carga el panel; combinado con el auto-refresh del panel admin,
    funciona como una sincronización 'casi en tiempo real' entre
    dispositivos (celulares que marcan y laptops que monitorean).

    Devuelve el DataFrame resultante (con las filas nuevas ya incluidas si
    las hubo) para que quien llama no tenga que releer el archivo del
    disco otra vez. Si Supabase no está disponible o falla, devuelve None
    y el que llama debe leer el CSV local por su cuenta.

    Optimización: cuando hay filas nuevas, se agregan al final del archivo
    (modo 'append') en vez de reescribir el CSV completo — con meses de
    historial esto es mucho más rápido y mantiene el bloqueo del archivo
    ocupado por menos tiempo."""
    if not supabase:
        return None
    try:
        with bloqueo_csv(CSV_ASISTENCIA):
            existe_archivo = os.path.exists(CSV_ASISTENCIA)
            if existe_archivo:
                df_local = pd.read_csv(CSV_ASISTENCIA)
            else:
                df_local = pd.DataFrame(columns=COLUMNAS_ASISTENCIA)

            desde = (hoy_peru() - timedelta(days=3)).strftime("%Y-%m-%d")
            res = (
                supabase.table("marcaciones_efimeras")
                .select("*")
                .eq("empresa_id", str(empresa_id))
                .gte("fecha", desde)
                .execute()
            )
            registros_nube = res.data or []
            if not registros_nube:
                st.session_state["_ultima_sync_hubo_cambios"] = False
                return df_local

            existentes = set()
            if not df_local.empty:
                for _, r in df_local.iterrows():
                    existentes.add((
                        str(r.get("Empleado", "")),
                        str(r.get("Fecha", "")),
                        str(r.get("Tipo Marcación", "")),
                        str(r.get("Hora Registrada", "")),
                    ))

            filas_nuevas = []
            for reg in registros_nube:
                clave = (
                    str(reg.get("nombre", "")),
                    str(reg.get("fecha", "")),
                    str(reg.get("tipo", "")),
                    str(reg.get("hora_registrada", "")),
                )
                if clave in existentes:
                    continue
                filas_nuevas.append({
                    "empresa_id": reg.get("empresa_id", empresa_id),
                    "Fecha": reg.get("fecha", ""),
                    "Empleado": reg.get("nombre", ""),
                    "Tipo Marcación": reg.get("tipo", ""),
                    "Hora Registrada": reg.get("hora_registrada", ""),
                    "Hora Entrada Oficial": reg.get(
                        "hora_entrada_oficial", ""
                    ),
                    "Hora Salida Oficial": reg.get(
                        "hora_salida_oficial", ""
                    ),
                    "Estado": reg.get("estado", ""),
                    "Minutos Tardanza": reg.get("minutos_tardanza", 0),
                    "Horas Extra (min)": reg.get("horas_extra_min", 0),
                    "Sede Detectada": reg.get("sede_detectada", ""),
                    "Distancia (m)": reg.get("distancia_m", 0.0),
                    "En Rango": reg.get("en_rango", ""),
                    "Foto": reg.get("foto_url", ""),
                })

            if not filas_nuevas:
                st.session_state["_ultima_sync_hubo_cambios"] = False
                return df_local

            df_nuevas = pd.DataFrame(filas_nuevas)[COLUMNAS_ASISTENCIA]
            df_nuevas.to_csv(
                CSV_ASISTENCIA,
                mode="a" if existe_archivo else "w",
                header=not existe_archivo,
                index=False,
            )
            st.session_state["_ultima_sync_hubo_cambios"] = True
            return pd.concat([df_local, df_nuevas], ignore_index=True)
    except Exception:
        return None  # si falla la sincronización, se lee el CSV local tal cual


def enviar_backup_email(asunto, cuerpo, adjuntos):
    """Envía un correo con el respaldo adjunto a la dirección del developer
    configurada en los Secrets de Streamlit. `adjuntos` es una lista de
    tuplas (nombre_archivo, bytes, tipo_mime)."""
    remitente = st.secrets.get("EMAIL_REMITENTE")
    clave_app = st.secrets.get("EMAIL_APP_PASSWORD")
    destino = st.secrets.get("EMAIL_DESTINO_DEVELOPER")

    if not remitente or not clave_app or not destino:
        raise RuntimeError(
            "Faltan los secrets EMAIL_REMITENTE, EMAIL_APP_PASSWORD o "
            "EMAIL_DESTINO_DEVELOPER en la configuración de Streamlit."
        )

    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = remitente
    msg["To"] = destino
    msg.set_content(cuerpo)

    for nombre, contenido, tipo_mime in adjuntos:
        maintype, subtype = tipo_mime.split("/", 1)
        msg.add_attachment(
            contenido, maintype=maintype, subtype=subtype, filename=nombre
        )

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(remitente, clave_app)
        server.send_message(msg)


def cargar_configuracion_sistema(supabase, empresa_id):
    """Carga PINs, clave de Excel y contraseña por defecto desde Supabase
    para esta empresa (tabla configuracion_sistema). Si no hay fila
    guardada todavía, deja los valores que ya estaban en session_state
    (los de Secrets o los de respaldo)."""
    if not supabase:
        return
    try:
        res = (
            supabase.table("configuracion_sistema")
            .select("*")
            .eq("empresa_id", str(empresa_id))
            .limit(1)
            .execute()
        )
        if res.data:
            cfg = res.data[0]
            if cfg.get("pin_admin"):
                st.session_state.pin_admin = cfg["pin_admin"]
            if cfg.get("pin_visor"):
                st.session_state.pin_visor = cfg["pin_visor"]
            if cfg.get("pin_master"):
                st.session_state.pin_master = cfg["pin_master"]
            if cfg.get("clave_excel"):
                st.session_state.clave_excel = cfg["clave_excel"]
    except Exception:
        pass  # si falla, se sigue usando lo que ya había cargado


def guardar_configuracion_sistema(supabase, empresa_id, **campos):
    """Guarda (crea o actualiza) los PINs/clave de esta empresa en
    Supabase, para que el cambio persista de verdad entre sesiones y
    redespliegues."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    datos = {"empresa_id": str(empresa_id), **campos}
    supabase.table("configuracion_sistema").upsert(
        datos, on_conflict="empresa_id"
    ).execute()


def obtener_password_empleado(supabase, empresa_id, dni, password_csv):
    """Devuelve la contraseña vigente de un trabajador: si tiene una
    contraseña propia guardada en Supabase (porque la cambió), usa esa;
    si no, usa la que viene del CSV local (la que le asignó el admin)."""
    if supabase:
        try:
            res = (
                supabase.table("empleados_passwords")
                .select("password")
                .eq("empresa_id", str(empresa_id))
                .eq("dni", str(dni))
                .limit(1)
                .execute()
            )
            if res.data:
                return res.data[0]["password"]
        except Exception as _e_silenciosa:
            logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")
    return password_csv


def guardar_password_empleado(supabase, empresa_id, dni, nueva_password):
    """Guarda la nueva contraseña que un trabajador eligió, de forma
    persistente en Supabase (para que sobreviva a los redespliegues)."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    datos = {
        "empresa_id": str(empresa_id),
        "dni": str(dni),
        "password": nueva_password,
    }
    supabase.table("empleados_passwords").upsert(
        datos, on_conflict="empresa_id,dni"
    ).execute()


# --- MIGRACIÓN: TABLA DE TRABAJADORES EN SUPABASE (antes vivía solo en un
# CSV local que Streamlit Cloud borra en cada redespliegue) ---
def cargar_empleados_supabase(supabase, empresa_id):
    """Trae la lista de trabajadores de esta empresa desde Supabase.
    Devuelve None si Supabase no está disponible o falla la consulta (así
    quien llama sabe que debe usar el CSV local como respaldo/modo offline).
    Devuelve una lista (puede estar vacía, [] ) si la consulta sí funcionó."""
    if not supabase:
        return None
    try:
        res = (
            supabase.table("empleados")
            .select("*")
            .eq("empresa_id", str(empresa_id))
            .execute()
        )
        return res.data
    except Exception:
        return None


def guardar_empleado_supabase(supabase, datos_empleado):
    """Crea o actualiza (upsert) un trabajador en Supabase. 'datos_empleado'
    debe incluir al menos empresa_id y dni; los demás campos que se pasen
    se sobrescriben, el resto de columnas de esa fila no se tocan."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    datos = dict(datos_empleado)
    datos["empresa_id"] = str(datos["empresa_id"])
    datos["dni"] = str(datos["dni"])
    supabase.table("empleados").upsert(
        datos, on_conflict="empresa_id,dni"
    ).execute()


def eliminar_empleado_supabase(supabase, empresa_id, dni):
    """Borra (DELETE real) un trabajador de Supabase."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    (
        supabase.table("empleados")
        .delete()
        .eq("empresa_id", str(empresa_id))
        .eq("dni", str(dni))
        .execute()
    )


# --- MIGRACIÓN: TABLA DE SEDES EN SUPABASE (mismo problema que empleados y
# empresas: antes vivía solo en un CSV local que se borra en cada reboot/
# redespliegue, perdiendo coordenadas GPS y horarios reales) ---
def cargar_sedes_supabase(supabase, empresa_id):
    """Trae las sedes de esta empresa desde Supabase. Devuelve None si
    Supabase no está disponible o falla la consulta."""
    if not supabase:
        return None
    try:
        res = (
            supabase.table("sedes")
            .select("*")
            .eq("empresa_id", str(empresa_id))
            .execute()
        )
        return res.data
    except Exception:
        return None


def guardar_sede_supabase(supabase, datos_sede):
    """Crea o actualiza (upsert) una sede en Supabase. 'datos_sede' debe
    incluir al menos empresa_id y nombre_sede."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    datos = dict(datos_sede)
    datos["empresa_id"] = str(datos["empresa_id"])
    datos["nombre_sede"] = str(datos["nombre_sede"])
    supabase.table("sedes").upsert(
        datos, on_conflict="empresa_id,nombre_sede"
    ).execute()


def eliminar_sede_supabase(supabase, empresa_id, nombre_sede):
    """Borra (DELETE real) una sede de Supabase."""
    if not supabase:
        raise RuntimeError("El cliente de Supabase no está configurado.")
    (
        supabase.table("sedes")
        .delete()
        .eq("empresa_id", str(empresa_id))
        .eq("nombre_sede", str(nombre_sede))
        .execute()
    )


def cargar_datos(empresa_id):
    cargar_empresas()

    registros_sedes = cargar_sedes_supabase(supabase, empresa_id)
    columnas_sedes = [
        "empresa_id",
        "nombre_sede",
        "latitud",
        "longitud",
        "hora_entrada",
        "hora_salida",
        "rango_metros",
    ]

    if registros_sedes is not None:
        # Supabase respondió: es la fuente de verdad (sobrevive a reboots
        # y redespliegues).
        if registros_sedes:
            df_sedes = pd.DataFrame(registros_sedes)
        else:
            df_sedes = pd.DataFrame(columns=columnas_sedes)

        if "empresa_id" not in df_sedes.columns:
            df_sedes["empresa_id"] = empresa_id
        if "rango_metros" not in df_sedes.columns:
            df_sedes["rango_metros"] = 100.0
        df_sedes["rango_metros"] = df_sedes["rango_metros"].fillna(100.0)
        if "hora_salida" not in df_sedes.columns:
            df_sedes["hora_salida"] = "17:00:00"
        df_sedes["hora_salida"] = df_sedes["hora_salida"].fillna("17:00:00")
        if "hora_entrada" not in df_sedes.columns:
            df_sedes["hora_entrada"] = "08:00:00"
        df_sedes["hora_entrada"] = df_sedes["hora_entrada"].fillna(
            "08:00:00"
        )

        # Copia local (caché/respaldo offline), preservando sedes de OTRAS
        # empresas que ya estuvieran en el CSV (panel Developer puede tener
        # varias empresas cargadas al ir cambiando de entorno).
        try:
            with bloqueo_csv(CSV_SEDES):
                if os.path.exists(CSV_SEDES):
                    df_csv_previo = pd.read_csv(CSV_SEDES)
                    df_csv_previo = df_csv_previo[
                        df_csv_previo["empresa_id"].astype(str)
                        != str(empresa_id)
                    ]
                    df_mirror = pd.concat(
                        [df_csv_previo, df_sedes[columnas_sedes]],
                        ignore_index=True,
                    )
                else:
                    df_mirror = df_sedes[columnas_sedes]
                df_mirror.to_csv(CSV_SEDES, index=False)
        except Exception as _e_silenciosa:
            logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")
        marcar_estado_modo_local("sedes", False)
    elif os.path.exists(CSV_SEDES):
        marcar_estado_modo_local("sedes", True)
        # Supabase no disponible ahora mismo: modo 100% local con lo
        # último que se guardó en el CSV.
        with bloqueo_csv(CSV_SEDES):
            df_sedes = pd.read_csv(CSV_SEDES)
            if "empresa_id" not in df_sedes.columns:
                df_sedes["empresa_id"] = empresa_id
            if "rango_metros" not in df_sedes.columns:
                df_sedes["rango_metros"] = 100.0
            if "hora_salida" not in df_sedes.columns:
                df_sedes["hora_salida"] = "17:00:00"
            df_sedes.to_csv(CSV_SEDES, index=False)
    else:
        marcar_estado_modo_local("sedes", True)
        # Ni Supabase ni CSV: primer arranque totalmente local, con sedes
        # de ejemplo para que la app no se caiga.
        df_sedes = pd.DataFrame({
            "empresa_id": [empresa_id] * 4,
            "nombre_sede": [
                "CASA ADMIN",
                "OFICINA PRINCIPAL",
                "TIENDA SUAREZ S&G",
                "TIENDA BALBOA INNOVACEROS",
            ],
            "latitud": [-8.0981, -8.1145, -8.1147, -8.1137],
            "longitud": [-79.0448, -79.0181, -79.0184, -79.0178],
            "hora_entrada": ["08:00:00", "08:00:00", "07:30:00", "07:30:00"],
            "hora_salida": ["17:00:00", "17:00:00", "16:30:00", "16:30:00"],
            "rango_metros": [100.0, 100.0, 100.0, 100.0],
        })
        with bloqueo_csv(CSV_SEDES):
            df_sedes.to_csv(CSV_SEDES, index=False)
        marcar_estado_modo_local("sedes", True)

    registros_empleados = cargar_empleados_supabase(supabase, empresa_id)

    if registros_empleados is not None:
        # Supabase respondió: es la fuente de verdad (sobrevive a los
        # redespliegues). Puede venir vacía si esta empresa todavía no
        # tiene trabajadores registrados en la nube.
        columnas_empleados = [
            "empresa_id",
            "dni",
            "nombre",
            "sede_principal",
            "sedes_autorizadas",
            "cargo",
            "password",
            "horario_personalizado",
            "fecha_ingreso",
        ]
        if registros_empleados:
            df_empleados = pd.DataFrame(registros_empleados)
        else:
            df_empleados = pd.DataFrame(columns=columnas_empleados)

        df_empleados["dni"] = df_empleados["dni"].astype(str)
        if "empresa_id" not in df_empleados.columns:
            df_empleados["empresa_id"] = empresa_id
        if "sedes_autorizadas" not in df_empleados.columns:
            df_empleados["sedes_autorizadas"] = "[]"
        df_empleados["sedes_autorizadas"] = df_empleados[
            "sedes_autorizadas"
        ].fillna("[]")
        if "password" not in df_empleados.columns:
            df_empleados["password"] = PASSWORD_EMPLEADO_DEFAULT
        df_empleados["password"] = df_empleados["password"].fillna(
            PASSWORD_EMPLEADO_DEFAULT
        )
        if "horario_personalizado" not in df_empleados.columns:
            df_empleados["horario_personalizado"] = "{}"
        df_empleados["horario_personalizado"] = df_empleados[
            "horario_personalizado"
        ].fillna("{}")
        if "fecha_ingreso" not in df_empleados.columns:
            df_empleados["fecha_ingreso"] = hoy_peru().strftime("%Y-%m-%d")
        df_empleados["fecha_ingreso"] = df_empleados["fecha_ingreso"].fillna(
            hoy_peru().strftime("%Y-%m-%d")
        )
        # Se guarda también una copia local, solo como caché/respaldo por
        # si más tarde Supabase no responde (modo offline de emergencia).
        # Se hace un "merge" con lo que ya había en el CSV para no perder
        # trabajadores de OTRAS empresas (el panel Developer puede tener
        # varias empresas cargadas en el mismo CSV al ir cambiando de
        # entorno/empresa desde la barra lateral).
        try:
            with bloqueo_csv(CSV_EMPLEADOS):
                if os.path.exists(CSV_EMPLEADOS):
                    df_csv_previo = pd.read_csv(CSV_EMPLEADOS)
                    df_csv_previo = df_csv_previo[
                        df_csv_previo["empresa_id"].astype(str)
                        != str(empresa_id)
                    ]
                    df_mirror = pd.concat(
                        [df_csv_previo, df_empleados[columnas_empleados]],
                        ignore_index=True,
                    )
                else:
                    df_mirror = df_empleados[columnas_empleados]
                df_mirror.to_csv(CSV_EMPLEADOS, index=False)
        except Exception as _e_silenciosa:
            logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")
        marcar_estado_modo_local("empleados", False)
    elif os.path.exists(CSV_EMPLEADOS):
        marcar_estado_modo_local("empleados", True)
        # Supabase no disponible ahora mismo: se sigue funcionando en modo
        # 100% local con lo último que se guardó en el CSV.
        with bloqueo_csv(CSV_EMPLEADOS):
            df_empleados = pd.read_csv(CSV_EMPLEADOS)
            df_empleados["dni"] = df_empleados["dni"].astype(str)

            if "empresa_id" not in df_empleados.columns:
                df_empleados["empresa_id"] = empresa_id
            if (
                "sede_asignada" in df_empleados.columns
                and "sede_principal" not in df_empleados.columns
            ):
                df_empleados.rename(
                    columns={"sede_asignada": "sede_principal"}, inplace=True
                )
            if "sedes_autorizadas" not in df_empleados.columns:
                df_empleados["sedes_autorizadas"] = df_empleados[
                    "sede_principal"
                ].apply(lambda x: json.dumps([x]) if pd.notna(x) else "[]")

            if "password" not in df_empleados.columns:
                df_empleados["password"] = PASSWORD_EMPLEADO_DEFAULT
            if "horario_personalizado" not in df_empleados.columns:
                df_empleados["horario_personalizado"] = "{}"
            if "fecha_ingreso" not in df_empleados.columns:
                df_empleados["fecha_ingreso"] = "2026-01-01"
            df_empleados.to_csv(CSV_EMPLEADOS, index=False)
    else:
        marcar_estado_modo_local("empleados", True)
        # Ni Supabase ni CSV: primer arranque totalmente local, con 2
        # trabajadores de ejemplo para que la app no se caiga.
        df_empleados = pd.DataFrame({
            "empresa_id": [empresa_id, empresa_id],
            "dni": ["75227702", "75227703"],
            "nombre": [
                "PEDRO FABIANO PAREDES RODRIGUEZ",
                "JUAN PEREZ ALVAREZ",
            ],
            "sede_principal": ["CASA ADMIN", "OFICINA PRINCIPAL"],
            "sedes_autorizadas": [
                json.dumps(["CASA ADMIN", "OFICINA PRINCIPAL"]),
                json.dumps(["OFICINA PRINCIPAL"]),
            ],
            "cargo": ["PRACTICANTE CONTABLE", "ASISTENTE DE VENTAS"],
            "password": [PASSWORD_EMPLEADO_DEFAULT, PASSWORD_EMPLEADO_DEFAULT],
            "horario_personalizado": ["{}", "{}"],
            "fecha_ingreso": ["2026-01-01", "2026-01-01"],
        })
        with bloqueo_csv(CSV_EMPLEADOS):
            df_empleados.to_csv(CSV_EMPLEADOS, index=False)

    df_asistencia_sync = sincronizar_marcaciones_nube(supabase, empresa_id)

    if df_asistencia_sync is not None:
        # La sincronización ya leyó (y si hizo falta, agregó) los datos;
        # se reutiliza en memoria sin volver a tocar el disco.
        df_asistencia = df_asistencia_sync
        columnas_faltantes = False
        if "empresa_id" not in df_asistencia.columns:
            df_asistencia["empresa_id"] = empresa_id
            columnas_faltantes = True
        if "Hora Salida Oficial" not in df_asistencia.columns:
            df_asistencia["Hora Salida Oficial"] = "17:00:00"
            columnas_faltantes = True
        if "Horas Extra (min)" not in df_asistencia.columns:
            df_asistencia["Horas Extra (min)"] = 0
            columnas_faltantes = True
        if columnas_faltantes:
            # Solo se reescribe el archivo si de verdad hubo que agregar
            # una columna nueva (por ejemplo, tras una actualización de
            # código) — no en cada carga de página normal.
            with bloqueo_csv(CSV_ASISTENCIA):
                df_asistencia.to_csv(CSV_ASISTENCIA, index=False)
    elif os.path.exists(CSV_ASISTENCIA):
        with bloqueo_csv(CSV_ASISTENCIA):
            df_asistencia = pd.read_csv(CSV_ASISTENCIA)
            columnas_faltantes = False
            if "empresa_id" not in df_asistencia.columns:
                df_asistencia["empresa_id"] = empresa_id
                columnas_faltantes = True
            if "Hora Salida Oficial" not in df_asistencia.columns:
                df_asistencia["Hora Salida Oficial"] = "17:00:00"
                columnas_faltantes = True
            if "Horas Extra (min)" not in df_asistencia.columns:
                df_asistencia["Horas Extra (min)"] = 0
                columnas_faltantes = True
            if columnas_faltantes:
                df_asistencia.to_csv(CSV_ASISTENCIA, index=False)
    else:
        df_asistencia = pd.DataFrame(columns=COLUMNAS_ASISTENCIA)
        with bloqueo_csv(CSV_ASISTENCIA):
            df_asistencia.to_csv(CSV_ASISTENCIA, index=False)

    # Asegurar filtrado robusto convirtiendo a string
    df_sedes_emp = df_sedes[df_sedes["empresa_id"].astype(str) == str(empresa_id)]
    df_empleados_emp = df_empleados[df_empleados["empresa_id"].astype(str) == str(empresa_id)]
    df_asistencia_emp = df_asistencia[df_asistencia["empresa_id"].astype(str) == str(empresa_id)]

    return df_sedes_emp, df_empleados_emp, df_asistencia_emp


# BARRA LATERAL: ENTORNO Y CAMBIO RÁPIDO
if not VISTA_TRABAJADOR_MOVIL:
    st.sidebar.title("📌 Menú Principal")

    if "dev_entorno_desbloqueado" not in st.session_state:
        st.session_state.dev_entorno_desbloqueado = False

    if not st.session_state.dev_entorno_desbloqueado:
        # Entorno DEV oculto para las empresas cliente: se fuerza PROD y
        # solo queda un candado discreto (sin texto explicativo) que pide
        # el PIN Developer para revelar el selector de entorno. La versión
        # celular no se toca — sigue igual que antes.
        if st.session_state.entorno != "PROD":
            st.session_state.entorno = "PROD"
        with st.sidebar.expander("🔒", expanded=False):
            _pin_candado_dev = st.text_input(
                "PIN",
                type="password",
                key="pin_candado_dev_input",
                label_visibility="collapsed",
            )
            if st.button("🔓", key="btn_candado_dev"):
                if clave_coincide(_pin_candado_dev, st.session_state.pin_master):
                    st.session_state.dev_entorno_desbloqueado = True
                    st.rerun()
                else:
                    st.error("PIN Incorrecto.")
    else:
        entorno_sel = st.sidebar.radio(
            "🌐 Entorno de Ejecución:",
            ["🚀 Producción", "🧪 Desarrollo / Sandbox"],
            index=0 if st.session_state.entorno == "PROD" else 1,
        )

        nuevo_entorno = "PROD" if entorno_sel == "🚀 Producción" else "DEV"

        if nuevo_entorno != st.session_state.entorno:
            st.session_state.entorno = nuevo_entorno
            df_emp_todas = cargar_empresas()
            empresas_filtradas = df_emp_todas[
                df_emp_todas["entorno"] == st.session_state.entorno
            ]
            if not empresas_filtradas.empty:
                st.session_state.empresa_id = empresas_filtradas.iloc[0]["empresa_id"]
            st.rerun()

        # Personalización de la animación de "globos" al marcar asistencia
        # (solo visible para el Developer con el entorno DEV desbloqueado).
        with st.sidebar.expander("🎈 Animación de éxito (solo dev)"):
            st.session_state.logo_globos_url = st.text_input(
                "URL de imagen para los globos:",
                value=st.session_state.get(
                    "logo_globos_url", "/app/static/icon-192.png"
                ),
                help=(
                    "Se usa en la animación que sube en globos al"
                    " confirmar una marcación. Por defecto es el ícono de"
                    " la app."
                ),
            )
elif EMPRESA_URL and not st.session_state.empresa_id:
    # En modo móvil, si la URL trae ?empresa=CODIGO, se precarga.
    st.session_state.empresa_id = EMPRESA_URL

df_empresas = cargar_empresas()
df_sedes, df_empleados, df_asistencia = cargar_datos(
    st.session_state.empresa_id
)

# Sincroniza el interruptor global de "Mejoras en Producción" con lo que
# haya guardado en Supabase, para que sea el mismo estado en todas las
# empresas, dispositivos y usuarios (ver funciones cerca de es_mejora_activa).
cargar_mejoras_prod_global(supabase)


# ---------------------------------------------------------
# 3. Funciones de Apoyo y Lógica
# ---------------------------------------------------------
def min_a_formato_horas(minutos_totales):
    hrs = int(minutos_totales // 60)
    mins = int(minutos_totales % 60)
    return f"{hrs}h {mins:02d}m"


def obtener_horario_oficial(emp_row, df_sedes, fecha_obj):
    nombre_dia = DIAS_SEMANA_MAP[fecha_obj.weekday()]
    horario_json = emp_row.get("horario_personalizado", "{}")
    if pd.notna(horario_json) and str(horario_json).strip() != "":
        try:
            h_dict = json.loads(str(horario_json))
            if nombre_dia in h_dict and h_dict[nombre_dia].get("activo", False):
                return (
                    h_dict[nombre_dia]["entrada"],
                    h_dict[nombre_dia]["salida"],
                )
        except Exception as _e_silenciosa:
            logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")

    sede_emp = emp_row["sede_principal"]
    datos_sede = df_sedes[df_sedes["nombre_sede"] == sede_emp]

    h_ent = (
        datos_sede["hora_entrada"].values[0]
        if not datos_sede.empty
        else "08:00:00"
    )
    h_sal = (
        datos_sede["hora_salida"].values[0]
        if not datos_sede.empty
        else "17:00:00"
    )
    return h_ent, h_sal


def calcular_distancia(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = np.radians(lat1), np.radians(lat2)
    delta_phi = np.radians(lat2 - lat1)
    delta_lambda = np.radians(lon2 - lon1)
    a = (
        np.sin(delta_phi / 2.0) ** 2
        + np.cos(phi1) * np.cos(phi2) * np.sin(delta_lambda / 2.0) ** 2
    )
    return R * (2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a)))


def validar_ubicacion(lat_user, lon_user, df_sedes, sedes_permitidas=None):
    if df_sedes.empty:
        return "Sin Sede", 0.0, "08:00:00", "17:00:00", 100.0

    df_eval = df_sedes
    if sedes_permitidas:
        df_eval = df_sedes[df_sedes["nombre_sede"].isin(sedes_permitidas)]
        if df_eval.empty:
            df_eval = df_sedes

    distancias = []
    for _, row in df_eval.iterrows():
        d = calcular_distancia(
            lat_user, lon_user, row["latitud"], row["longitud"]
        )
        rango = row["rango_metros"] if "rango_metros" in row else 100.0
        h_salida = row["hora_salida"] if "hora_salida" in row else "17:00:00"
        distancias.append((
            row["nombre_sede"],
            d,
            row["hora_entrada"],
            h_salida,
            rango,
        ))
    distancias.sort(key=lambda x: x[1])
    return distancias[0]


@st.cache_data(show_spinner=False, ttl=3600)
def _descargar_foto_storage(nombre_foto):
    """Descarga una foto desde Supabase Storage (bucket fotos-asistencia)
    y la deja en caché un rato, para no repetir la descarga en cada
    autorefresh del panel. Devuelve None si no se pudo conseguir."""
    if not supabase or not nombre_foto:
        return None
    try:
        return supabase.storage.from_("fotos-asistencia").download(
            os.path.basename(str(nombre_foto))
        )
    except Exception:
        return None


def render_avatar_with_zoom(foto_path, index_id):
    contenido_foto = None

    if foto_path and not pd.isna(foto_path):
        if os.path.exists(str(foto_path)):
            try:
                with open(str(foto_path), "rb") as image_file:
                    contenido_foto = image_file.read()
            except Exception:
                contenido_foto = None

        if contenido_foto is None:
            # El archivo ya no está en el disco local (se borra en cada
            # reboot/redespliegue) — se intenta traer desde Supabase
            # Storage, donde sí queda guardado de forma permanente.
            contenido_foto = _descargar_foto_storage(foto_path)

    if not contenido_foto:
        return "<span style='color: #6c757d;'>—</span>"

    try:
        encoded_string = base64.b64encode(contenido_foto).decode("utf-8")

        img_src = f"data:image/png;base64,{encoded_string}"
        modal_id = f"img-modal-{index_id}"

        return (
            f'<a href="#{modal_id}"><img src="{img_src}"'
            ' class="user-avatar-thumb" title="Clic para ampliar" /></a><div'
            f' id="{modal_id}" class="img-modal-backdrop"><a href="#_"'
            ' class="img-modal-close-overlay"></a><div'
            ' class="img-modal-content"><a href="#_" class="img-modal-close"'
            f' title="Cerrar (ESC)">✕</a><img src="{img_src}" /></div></div>'
        )
    except Exception:
        return "<span style='color: #6c757d;'>—</span>"


def render_custom_table(lista_registros):
    html_lines = [
        '<div class="bitacora-container"><table class="bitacora-table">',
        "<thead><tr>",
        "<th>Fecha / Día</th>",
        "<th>Tipo Marcación</th>",
        "<th>Hora Registrada</th>",
        "<th>Estado</th>",
        "<th>Min. Tardanza</th>",
        "<th>Horas Extra (min)</th>",
        "<th>Sede Detectada</th>",
        "<th>Distancia (m)</th>",
        "<th>En Rango</th>",
        '<th style="text-align:center;">Foto Evidencia</th>',
        "</tr></thead><tbody>",
    ]

    previous_week = None

    for idx, reg in enumerate(lista_registros):
        fecha_raw = reg.get("Fecha_Raw")
        current_week = None
        if fecha_raw:
            try:
                fecha_dt = pd.to_datetime(fecha_raw)
                current_week = fecha_dt.isocalendar()[1]
            except Exception as _e_silenciosa:
                logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")

        is_new_week = (
            previous_week is not None
            and current_week is not None
            and current_week != previous_week
        )
        row_class = ' class="week-divider"' if is_new_week else ""
        if current_week is not None:
            previous_week = current_week

        est_raw = str(reg.get("Estado", "")).upper()
        if "PUNTUAL" in est_raw:
            estado_html = (
                f'<span class="badge-puntual">● {reg["Estado"]}</span>'
            )
        elif "TARDANZA" in est_raw:
            estado_html = (
                f'<span class="badge-tardanza">● {reg["Estado"]}</span>'
            )
        elif "FALTA" in est_raw:
            estado_html = f'<span class="badge-falta">● {reg["Estado"]}</span>'
        elif "FERIADO" in est_raw:
            estado_html = (
                f'<span class="badge-feriado">● {reg["Estado"]}</span>'
            )
        else:
            estado_html = reg.get("Estado", "")

        foto_html = render_avatar_with_zoom(reg.get("Foto"), idx)
        dist_text = (
            f"{reg.get('Distancia (m)')} m"
            if reg.get("Distancia (m)") != "—"
            else "—"
        )

        row_html = (
            f"<tr{row_class}>"
            f"<td><b>{reg.get('Fecha_Display', '')}</b></td>"
            f"<td>{reg.get('Tipo Marcación', '')}</td>"
            f"<td>{reg.get('Hora Registrada', '')}</td>"
            f"<td>{estado_html}</td>"
            f"<td>{reg.get('Minutos Tardanza', 0)} min</td>"
            f"<td>{reg.get('Horas Extra (min)', 0)} min</td>"
            f"<td>{reg.get('Sede Detectada', '')}</td>"
            f"<td>{dist_text}</td>"
            f"<td>{reg.get('En Rango', '')}</td>"
            f'<td style="text-align:center;">{foto_html}</td>'
            f"</tr>"
        )
        html_lines.append(row_html)

    html_lines.append("</tbody></table></div>")
    return "".join(html_lines)


def generar_excel_completo(
    df_asistencia, df_empleados, mes_sel, anio_sel, clave_excel
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_navy = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")

    fill_rojo = PatternFill(
        start_color="FF4D4D", end_color="FF4D4D", fill_type="solid"
    )
    fill_naranja = PatternFill(
        start_color="FF8C00", end_color="FF8C00", fill_type="solid"
    )
    fill_verde = PatternFill(
        start_color="2EB67D", end_color="2EB67D", fill_type="solid"
    )
    font_blanca = Font(name="Calibri", size=10, color="FFFFFF", bold=True)

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    prefix_periodo = f"{anio_sel}-{mes_sel:02d}"

    ws_gen = wb.create_sheet(title="Resumen General")
    ws_gen.append(["REPORTE CONSOLIDADO DE ASISTENCIA Y AUDITORÍA GPS"])
    ws_gen.cell(row=1, column=1).font = font_title
    ws_gen.append([f"Período Evaluado: {MESES_NOMBRES[mes_sel]} {anio_sel}"])
    ws_gen.append([])

    headers_gen = [
        "DNI",
        "Empleado",
        "Sede Principal",
        "Cargo",
        "Días Puntuales",
        "Tardanzas",
        "Horas Tardanza",
        "Minutos Extra",
    ]
    ws_gen.append(headers_gen)
    for col_idx in range(1, len(headers_gen) + 1):
        c = ws_gen.cell(row=4, column=col_idx)
        c.fill, c.font, c.alignment = (
            fill_navy,
            font_header,
            Alignment(horizontal="center", vertical="center"),
        )

    row_pos = 5
    for _, emp in df_empleados.iterrows():
        emp_asist = df_asistencia[(df_asistencia["Empleado"] == emp["nombre"])]
        emp_asist_mes = (
            emp_asist[
                emp_asist["Fecha"].astype(str).str.startswith(prefix_periodo)
            ]
            if not emp_asist.empty
            else pd.DataFrame()
        )

        tardanzas = (
            emp_asist_mes[emp_asist_mes["Estado"] == "Tardanza"][
                "Fecha"
            ].nunique()
            if not emp_asist_mes.empty
            else 0
        )
        puntuales = (
            emp_asist_mes[emp_asist_mes["Estado"] == "Puntual"][
                "Fecha"
            ].nunique()
            if not emp_asist_mes.empty
            else 0
        )
        min_tard = (
            emp_asist_mes["Minutos Tardanza"].sum()
            if not emp_asist_mes.empty
            else 0
        )
        min_extra = (
            emp_asist_mes["Horas Extra (min)"].sum()
            if not emp_asist_mes.empty
            else 0
        )

        hrs_tard = round(min_tard / 60.0, 2)

        ws_gen.append([
            emp["dni"],
            emp["nombre"],
            emp["sede_principal"],
            emp["cargo"],
            puntuales,
            tardanzas,
            hrs_tard,
            min_extra,
        ])
        for c_i in range(1, len(headers_gen) + 1):
            cell = ws_gen.cell(row=row_pos, column=c_i)
            cell.border = border_thin
            cell.alignment = Alignment(
                horizontal="center" if c_i != 2 else "left", vertical="center"
            )
        row_pos += 1

    for _, emp in df_empleados.iterrows():
        nombre_sheet = str(emp["nombre"]).split()[0] + f"_{emp['dni'][-4:]}"
        ws_emp = wb.create_sheet(title=nombre_sheet[:30])

        ws_emp.append([f"FICHA INDIVIDUAL DE CONTROL: {emp['nombre']}"])
        ws_emp.cell(row=1, column=1).font = font_title
        ws_emp.append([
            f"DNI: {emp['dni']} | Cargo: {emp['cargo']} | Sede Principal:"
            f" {emp['sede_principal']} | Mes: {MESES_NOMBRES[mes_sel]} {anio_sel}"
        ])
        ws_emp.append([])

        cols_emp = [
            "Fecha",
            "Marcación",
            "Hora Reg.",
            "Hora Entrada Ofic.",
            "Estado",
            "Min. Tardanza",
            "Min. Extra",
            "GPS Sede",
            "Distancia (m)",
            "En Rango",
            "Foto Evidencia",
        ]
        ws_emp.append(cols_emp)

        for c_idx in range(1, len(cols_emp) + 1):
            cell = ws_emp.cell(row=4, column=c_idx)
            cell.fill, cell.font, cell.alignment = (
                fill_navy,
                font_header,
                Alignment(horizontal="center", vertical="center"),
            )

        df_registros = df_asistencia[
            (df_asistencia["Empleado"] == emp["nombre"])
            & (df_asistencia["Fecha"].astype(str).str.startswith(prefix_periodo))
        ]

        r_idx = 5
        for _, reg in df_registros.iterrows():
            ws_emp.append([
                reg["Fecha"],
                reg["Tipo Marcación"],
                reg["Hora Registrada"],
                reg["Hora Entrada Oficial"],
                reg["Estado"],
                reg["Minutos Tardanza"],
                reg.get("Horas Extra (min)", 0),
                reg["Sede Detectada"],
                reg["Distancia (m)"],
                reg["En Rango"],
                "",
            ])

            ws_emp.row_dimensions[r_idx].height = 40
            for c_i in range(1, len(cols_emp) + 1):
                cell = ws_emp.cell(row=r_idx, column=c_i)
                cell.border = border_thin
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                if c_i == 5:
                    est = str(reg["Estado"]).upper()
                    if "TARDANZA" in est:
                        cell.fill, cell.font = fill_naranja, font_blanca
                    elif "PUNTUAL" in est:
                        cell.fill, cell.font = fill_verde, font_blanca
                    else:
                        cell.fill, cell.font = fill_rojo, font_blanca

            foto_p = reg.get("Foto", "")
            if pd.notna(foto_p):
                try:
                    if os.path.exists(str(foto_p)):
                        img = OpenPyxlImage(str(foto_p))
                    else:
                        # No está en el disco local (se borra en cada
                        # reboot/redespliegue): se intenta traer desde
                        # Supabase Storage, donde queda guardada siempre.
                        contenido_foto_xl = _descargar_foto_storage(foto_p)
                        img = (
                            OpenPyxlImage(io.BytesIO(contenido_foto_xl))
                            if contenido_foto_xl
                            else None
                        )
                    if img is not None:
                        img.width, img.height = 55, 38
                        col_let = get_column_letter(11)
                        img.anchor = f"{col_let}{r_idx}"
                        ws_emp.add_image(img)
                except Exception as _e_silenciosa:
                    logger.warning(f"Error controlado (ignorado para el usuario): {_e_silenciosa}")

            r_idx += 1

        for col in ws_emp.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_emp.column_dimensions[get_column_letter(col[0].column)].width = (
                max(max_len + 3, 13)
            )

        ws_emp.protection.sheet = True
        ws_emp.protection.set_password(clave_excel)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def render_modulo_sedes(df_sedes):
    col_s1, col_s2 = st.columns([2, 1.2])
    with col_s1:
        st.dataframe(df_sedes, use_container_width=True, hide_index=True)

    with col_s2:
        st.markdown("#### 📝 Crear / Editar Sede")

        opciones_sedes = ["-- Crear Nueva Sede --"] + list(
            df_sedes["nombre_sede"].unique()
        )
        sede_sel_ed = st.selectbox(
            "Seleccionar Sede para Editar:", opciones_sedes
        )

        is_edit_s = sede_sel_ed != "-- Crear Nueva Sede --"

        if is_edit_s:
            datos_s = df_sedes[df_sedes["nombre_sede"] == sede_sel_ed].iloc[0]
            val_s_nombre = str(datos_s["nombre_sede"])
            val_s_lat = float(datos_s["latitud"])
            val_s_lon = float(datos_s["longitud"])
            val_s_ent = datetime.strptime(
                str(datos_s["hora_entrada"]), "%H:%M:%S"
            ).time()
            val_s_sal = datetime.strptime(
                str(datos_s["hora_salida"]), "%H:%M:%S"
            ).time()
            val_s_rango = float(datos_s.get("rango_metros", 100.0))
        else:
            val_s_nombre = ""
            val_s_lat = -8.098100
            val_s_lon = -79.044800
            val_s_ent = time(8, 0)
            val_s_sal = time(17, 0)
            val_s_rango = 100.0

        nueva_s_nombre = st.text_input(
            "Nombre de Sede:",
            value=val_s_nombre,
            disabled=is_edit_s,
        )
        nueva_s_lat = st.number_input(
            "Latitud:", format="%.6f", value=val_s_lat
        )
        nueva_s_lon = st.number_input(
            "Longitud:", format="%.6f", value=val_s_lon
        )
        nueva_s_ent = st.time_input("Hora Entrada:", value=val_s_ent)
        nueva_s_sal = st.time_input("Hora Salida:", value=val_s_sal)
        nueva_s_rango = st.number_input("Radio Máximo (m):", value=val_s_rango)

        col_btn_s1, col_btn_s2 = st.columns(2)

        with col_btn_s1:
            if is_edit_s:
                if st.button("💾 Actualizar Sede", use_container_width=True):
                    datos_sede_upd = {
                        "empresa_id": st.session_state.empresa_id,
                        "nombre_sede": sede_sel_ed,
                        "latitud": nueva_s_lat,
                        "longitud": nueva_s_lon,
                        "hora_entrada": nueva_s_ent.strftime("%H:%M:%S"),
                        "hora_salida": nueva_s_sal.strftime("%H:%M:%S"),
                        "rango_metros": nueva_s_rango,
                    }
                    if supabase:
                        try:
                            guardar_sede_supabase(supabase, datos_sede_upd)
                        except Exception as e:
                            st.warning(
                                "No se pudo guardar en la nube"
                                f" ({e}). Se guardó solo local; se"
                                " perderá en el próximo redespliegue."
                            )

                    if os.path.exists(CSV_SEDES):
                        with bloqueo_csv(CSV_SEDES):
                            df_sedes_full = pd.read_csv(CSV_SEDES)
                            idx_s = df_sedes_full[
                                (
                                    df_sedes_full["empresa_id"].astype(str)
                                    == str(st.session_state.empresa_id)
                                )
                                & (df_sedes_full["nombre_sede"] == sede_sel_ed)
                            ].index
                            if len(idx_s) > 0:
                                for campo, valor in datos_sede_upd.items():
                                    if campo not in (
                                        "empresa_id",
                                        "nombre_sede",
                                    ):
                                        df_sedes_full.at[
                                            idx_s[0], campo
                                        ] = valor
                                df_sedes_full.to_csv(CSV_SEDES, index=False)
                    st.success("Sede actualizada correctamente.")
                    st.rerun()
            else:
                if st.button("➕ Crear Sede", use_container_width=True):
                    if nueva_s_nombre:
                        nueva_fila = {
                            "empresa_id": st.session_state.empresa_id,
                            "nombre_sede": nueva_s_nombre.strip().upper(),
                            "latitud": nueva_s_lat,
                            "longitud": nueva_s_lon,
                            "hora_entrada": nueva_s_ent.strftime("%H:%M:%S"),
                            "hora_salida": nueva_s_sal.strftime("%H:%M:%S"),
                            "rango_metros": nueva_s_rango,
                        }

                        if supabase:
                            try:
                                guardar_sede_supabase(supabase, nueva_fila)
                            except Exception as e:
                                st.warning(
                                    "No se pudo guardar en la nube"
                                    f" ({e}). Se guardó solo local; se"
                                    " perderá en el próximo redespliegue."
                                )

                        if os.path.exists(CSV_SEDES):
                            with bloqueo_csv(CSV_SEDES):
                                df_sedes_full = pd.read_csv(CSV_SEDES)
                                df_sedes_full = pd.concat(
                                    [df_sedes_full, pd.DataFrame([nueva_fila])],
                                    ignore_index=True,
                                )
                                df_sedes_full.to_csv(CSV_SEDES, index=False)
                        else:
                            df_sedes_full = pd.DataFrame([nueva_fila])
                            df_sedes_full.to_csv(CSV_SEDES, index=False)
                        st.success(f"Sede {nueva_s_nombre} creada con éxito.")
                        st.rerun()

        with col_btn_s2:
            if is_edit_s:
                if st.button("🗑️ Eliminar Sede", use_container_width=True):
                    if supabase:
                        try:
                            eliminar_sede_supabase(
                                supabase,
                                st.session_state.empresa_id,
                                sede_sel_ed,
                            )
                        except Exception as e:
                            st.warning(
                                "No se pudo eliminar en la nube"
                                f" ({e}). Se eliminó solo local."
                            )
                    if os.path.exists(CSV_SEDES):
                        with bloqueo_csv(CSV_SEDES):
                            df_sedes_full = pd.read_csv(CSV_SEDES)
                            df_sedes_full = df_sedes_full[
                                ~(
                                    (
                                        df_sedes_full["empresa_id"].astype(str)
                                        == str(st.session_state.empresa_id)
                                    )
                                    & (
                                        df_sedes_full["nombre_sede"]
                                        == sede_sel_ed
                                    )
                                )
                            ]
                            df_sedes_full.to_csv(CSV_SEDES, index=False)
                    st.warning("Sede eliminada.")
                    st.rerun()


def render_modulo_empresas():
    st.markdown("#### 🏢 Administración y Edición de Empresas")

    c_dev1, c_dev2 = st.columns([2.2, 1.3])

    df_emp_list = cargar_empresas()

    with c_dev1:
        st.markdown("##### 📋 Listado Registrado")
        if os.path.exists(CSV_EMPLEADOS):
            df_all_emp = pd.read_csv(CSV_EMPLEADOS)
            conteo = df_all_emp.groupby("empresa_id")["dni"].count().to_dict()
            df_emp_list["Trabajadores"] = df_emp_list["empresa_id"].map(
                lambda x: conteo.get(x, 0)
            )

        st.dataframe(df_emp_list, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("##### 🚀 Activar Mejoras en Producción")
        st.caption(
            "Habilita las nuevas funciones desarrolladas para todas las"
            " empresas en PRODUCCIÓN. Las empresas en DEV siempre las tienen"
            " activas."
        )

        valido, msg = validar_integridad_desarrollo()
        estado_actual = st.session_state.get("mejoras_activadas_prod", False)

        if estado_actual:
            st.info("🟢 Estado: Las nuevas mejoras están ACTIVAS permanentemente en Producción.")
        else:
            if valido:
                st.warning("🟡 Estado: Las nuevas mejoras solo están activas en DEV (Entorno listo para despliegue).")
            else:
                st.error(f"🔴 Estado: Entorno DEV no verificado ({msg}).")

            btn_activar = st.button(
                "⚡ Activar y Desplegar Mejoras a Producción",
                type="primary",
                use_container_width=True,
                key="btn_activar_prod"
            )
            
            if btn_activar:
                modal_confirmar_despliegue()

    with c_dev2:
        st.markdown("##### 📝 Crear / Editar Empresa")

        opciones_empresas = ["-- Registrar Nueva Empresa --"] + list(
            df_emp_list["empresa_id"].unique()
        )
        emp_sel_ed = st.selectbox("Seleccionar Empresa:", opciones_empresas)

        is_edit_emp = emp_sel_ed != "-- Registrar Nueva Empresa --"

        if is_edit_emp:
            datos_emp_ed = df_emp_list[
                df_emp_list["empresa_id"] == emp_sel_ed
            ].iloc[0]
            v_emp_code = str(datos_emp_ed["empresa_id"])
            v_emp_rz = str(datos_emp_ed["razon_social"])
            v_emp_ruc = str(datos_emp_ed.get("ruc", ""))
            v_emp_plan = str(datos_emp_ed.get("plan", "BASIC"))
            v_emp_ent = str(datos_emp_ed.get("entorno", "PROD"))
        else:
            v_emp_code = ""
            v_emp_rz = ""
            v_emp_ruc = ""
            v_emp_plan = "BASIC"
            v_emp_ent = "PROD"

        ed_code = st.text_input(
            "Código Empresa (ID):",
            value=v_emp_code,
            disabled=is_edit_emp,
        )
        ed_rz = st.text_input("Razón Social / Nombre:", value=v_emp_rz)
        ed_ruc = st.text_input("RUC:", value=v_emp_ruc)
        ed_plan = st.selectbox(
            "Plan SaaS:",
            ["BASIC", "PREMIUM", "ENTERPRISE", "DEVELOPER"],
            index=(
                ["BASIC", "PREMIUM", "ENTERPRISE", "DEVELOPER"].index(
                    v_emp_plan
                )
                if v_emp_plan in ["BASIC", "PREMIUM", "ENTERPRISE", "DEVELOPER"]
                else 0
            ),
        )
        ed_entorno = st.selectbox(
            "Entorno:", ["PROD", "DEV"], index=0 if v_emp_ent == "PROD" else 1
        )

        col_eb1, col_eb2 = st.columns(2)

        with col_eb1:
            if is_edit_emp:
                if st.button("💾 Guardar Empresa", use_container_width=True):
                    datos_emp_upd = {
                        "empresa_id": emp_sel_ed,
                        "razon_social": ed_rz.strip().upper(),
                        "ruc": ed_ruc.strip(),
                        "plan": ed_plan,
                        "entorno": ed_entorno,
                    }
                    if supabase:
                        try:
                            guardar_empresa_supabase(supabase, datos_emp_upd)
                        except Exception as e:
                            st.warning(
                                "No se pudo guardar en la nube"
                                f" ({e}). Se guardó solo local; se"
                                " perderá en el próximo redespliegue."
                            )

                    df_e_all = cargar_empresas()
                    idx_e = df_e_all[
                        df_e_all["empresa_id"] == emp_sel_ed
                    ].index
                    if len(idx_e) > 0:
                        for campo, valor in datos_emp_upd.items():
                            if campo != "empresa_id":
                                df_e_all.at[idx_e[0], campo] = valor
                        with bloqueo_csv(CSV_EMPRESAS):
                            df_e_all.to_csv(CSV_EMPRESAS, index=False)
                    st.success("Datos de la empresa guardados.")
                    st.rerun()
            else:
                if st.button("➕ Crear Empresa", use_container_width=True):
                    if ed_code and ed_rz:
                        code_c = ed_code.strip().upper()
                        df_e_all = cargar_empresas()

                        if code_c in df_e_all["empresa_id"].values:
                            st.error("El código de empresa ya existe.")
                        else:
                            new_e = {
                                "empresa_id": code_c,
                                "razon_social": ed_rz.strip().upper(),
                                "ruc": ed_ruc.strip(),
                                "plan": ed_plan,
                                "estado": "ACTIVO",
                                "entorno": ed_entorno,
                            }

                            if supabase:
                                try:
                                    guardar_empresa_supabase(
                                        supabase, new_e
                                    )
                                except Exception as e:
                                    st.warning(
                                        "No se pudo guardar en la nube"
                                        f" ({e}). Se guardó solo local;"
                                        " se perderá en el próximo"
                                        " redespliegue."
                                    )

                            df_e_all = pd.concat(
                                [df_e_all, pd.DataFrame([new_e])],
                                ignore_index=True,
                            )
                            with bloqueo_csv(CSV_EMPRESAS):
                                df_e_all.to_csv(CSV_EMPRESAS, index=False)
                            st.success(f"Empresa '{code_c}' creada.")
                            st.rerun()

        with col_eb2:
            if is_edit_emp:
                if st.button("🗑️ Eliminar Empresa", use_container_width=True):
                    if supabase:
                        try:
                            eliminar_empresa_supabase(supabase, emp_sel_ed)
                        except Exception as e:
                            st.warning(
                                "No se pudo eliminar en la nube"
                                f" ({e}). Se eliminó solo local."
                            )
                    df_e_all = cargar_empresas()
                    df_e_all = df_e_all[df_e_all["empresa_id"] != emp_sel_ed]
                    with bloqueo_csv(CSV_EMPRESAS):
                        df_e_all.to_csv(CSV_EMPRESAS, index=False)
                    st.warning("Empresa eliminada.")
                    st.rerun()


# ---------------------------------------------------------
# 4. Interfaz Visual
# ---------------------------------------------------------

if st.session_state.emp_login_ok and not st.session_state.autenticado:
    st.markdown(
        """
        <style>
            [data-testid="stSidebar"] {display: none;}
        </style>
    """,
        unsafe_allow_html=True,
    )

if st.session_state.entorno == "DEV" and not VISTA_TRABAJADOR_MOVIL:
    st.sidebar.warning(
        "⚠️ Entorno de Desarrollo Activo (Solo empresas Sandbox)"
    )

if VISTA_TRABAJADOR_MOVIL:
    opcion = "⏰ Marcar Asistencia"
else:
    opcion = st.sidebar.radio(
        "Ir a:", ["⏰ Marcar Asistencia", "🔐 Panel de Gestión / Admin"]
    )

    if st.sidebar.button("🚪 Cerrar Sesión"):
        st.session_state.autenticado = False
        st.session_state.rol = None
        st.session_state.emp_login_ok = False
        st.session_state.emp_datos = None
        st.session_state.dev_entorno_desbloqueado = False
        st.session_state.entorno = "PROD"
        st.rerun()

if opcion == "⏰ Marcar Asistencia":
    st.markdown(
        f"""
        <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:6px;">
            <div style="display:flex; align-items:center; gap:10px;">
                <div style="width:38px; height:38px; border-radius:11px;
                    background:linear-gradient(135deg, var(--cyan), var(--violet));
                    display:flex; align-items:center; justify-content:center;
                    font-size:18px; box-shadow:0 0 22px rgba(88,166,255,0.35);">⏰</div>
                <div style="font-family:var(--font-display); font-weight:700; font-size:17px;">
                    Registro de Asistencia
                    <small style="display:block; font-weight:400; font-size:11px; color:var(--text-muted);">
                        Marcación por GPS + foto
                    </small>
                </div>
            </div>
            <div style="text-align:right;">
                <div style="font-family:var(--font-mono); font-weight:700; font-size:22px;
                    background:linear-gradient(90deg, #fff, var(--cyan) 120%);
                    -webkit-background-clip:text; background-clip:text; color:transparent;">
                    {ahora_peru().strftime('%H:%M')}
                </div>
                <div style="font-size:11.5px; color:var(--text-muted);">
                    {ahora_peru().strftime('%d/%m/%Y')}
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    hoy = hoy_peru()

    if VISTA_TRABAJADOR_MOVIL:
        with st.expander("🔐 ¿Eres Admin, SuperAdmin o Developer?"):
            df_e_disponibles_mov = cargar_empresas()
            df_e_disponibles_mov = df_e_disponibles_mov[
                df_e_disponibles_mov["entorno"] == st.session_state.entorno
            ]
            empresa_admin_mov = st.selectbox(
                "Empresa:",
                df_e_disponibles_mov["empresa_id"].unique()
                if not df_e_disponibles_mov.empty
                else [],
                key="empresa_admin_movil",
            )
            if empresa_admin_mov:
                cargar_configuracion_sistema(supabase, empresa_admin_mov)
            pin_mov = st.text_input(
                "PIN de Acceso:", type="password", key="pin_admin_movil"
            )
            if st.button("Ingresar al Panel", key="btn_login_admin_movil"):
                if empresa_admin_mov:
                    st.session_state.empresa_id = empresa_admin_mov
                    if clave_coincide(pin_mov, st.session_state.pin_admin):
                        st.session_state.autenticado = True
                        st.session_state.rol = "admin"
                        st.rerun()
                    elif clave_coincide(pin_mov, st.session_state.pin_visor):
                        st.session_state.autenticado = True
                        st.session_state.rol = "visor"
                        st.rerun()
                    elif clave_coincide(pin_mov, st.session_state.pin_master):
                        st.session_state.autenticado = True
                        st.session_state.rol = "master"
                        st.rerun()
                    else:
                        st.error("PIN Incorrecto.")
                else:
                    st.error("No hay empresas disponibles en este entorno.")

    if not st.session_state.emp_login_ok:
        st.markdown("### 🔑 Iniciar Sesión de Empleado")
        col_acc1, col_acc2 = st.columns(2)
        with col_acc1:
            # Selector desplegable de empresa (antes era un campo de texto
            # libre). Se filtra SOLO por el modo en el que esté el usuario:
            # si no inició sesión como Developer (PIN 9999), solo ve las
            # empresas de PRODUCCIÓN; si sí inició sesión como Developer,
            # solo ve las de DESARROLLO/Sandbox. Aplica igual en celular y
            # en PC — no depende del toggle abierto de entorno del sidebar.
            _dev_autenticado = st.session_state.get("rol") == "master"
            _entorno_login_empleado = "DEV" if _dev_autenticado else "PROD"

            df_e_disponibles_login = cargar_empresas()
            df_e_disponibles_login = df_e_disponibles_login[
                df_e_disponibles_login["entorno"] == _entorno_login_empleado
            ]
            opciones_empresa_login = (
                list(df_e_disponibles_login["empresa_id"].astype(str).unique())
                if not df_e_disponibles_login.empty
                else []
            )

            if opciones_empresa_login:
                idx_empresa_login = (
                    opciones_empresa_login.index(st.session_state.empresa_id)
                    if st.session_state.empresa_id in opciones_empresa_login
                    else 0
                )
                empresa_input = st.selectbox(
                    "Código de Empresa:",
                    opciones_empresa_login,
                    index=idx_empresa_login,
                )
            else:
                empresa_input = None
                st.info(
                    "No hay empresas configuradas todavía en este entorno."
                )
            dni_input = st.text_input("Ingrese su DNI:")
            pass_input = st.text_input(
                "Ingrese su Contraseña:", type="password"
            )

            if st.button("Ingresar para Marcar"):
                if dni_input and pass_input and empresa_input:
                    emp_code = empresa_input.strip().upper()
                    df_e_val = cargar_empresas()
                    emp_find = df_e_val[df_e_val["empresa_id"] == emp_code]

                    if (
                        emp_find.empty
                        or emp_find.iloc[0]["entorno"]
                        != st.session_state.entorno
                    ):
                        st.error(
                            f"La empresa '{emp_code}' no existe o no pertenece"
                            f" al entorno {st.session_state.entorno}."
                        )
                    else:
                        st.session_state.empresa_id = emp_code
                        (
                            df_sedes_emp,
                            df_empleados_emp,
                            df_asistencia_emp,
                        ) = cargar_datos(st.session_state.empresa_id)

                        emp_match = df_empleados_emp[
                            df_empleados_emp["dni"].astype(str)
                            == dni_input.strip()
                        ]
                        login_ok = False
                        if not emp_match.empty:
                            fila_emp = emp_match.iloc[0]
                            password_vigente = obtener_password_empleado(
                                supabase,
                                st.session_state.empresa_id,
                                fila_emp["dni"],
                                str(fila_emp["password"]),
                            )
                            login_ok = clave_coincide(
                                pass_input.strip(), password_vigente
                            )
                        if login_ok:
                            st.session_state.emp_login_ok = True
                            st.session_state.emp_datos = emp_match.iloc[0]
                            st.success("Acceso verificado correctamente.")
                            st.rerun()
                        else:
                            st.error(
                                "DNI o Contraseña incorrectos para esta"
                                " empresa."
                            )
                else:
                    st.warning("Por favor complete todos los campos.")
    else:
        datos_emp = st.session_state.emp_datos

        col_top1, col_top2 = st.columns([3, 1])
        with col_top1:
            st.success(
                f"👤 Bienvenid@, **{datos_emp['nombre']}** ({datos_emp['cargo']})"
            )
        with col_top2:
            if st.button("🔒 Salir de Marcación"):
                st.session_state.emp_login_ok = False
                st.session_state.emp_datos = None
                st.rerun()

        st.divider()

        with st.expander("🔑 Cambiar mi Contraseña"):
            pass_actual = st.text_input(
                "Contraseña actual:", type="password", key="cambio_pass_actual"
            )
            pass_nueva = st.text_input(
                "Nueva contraseña:", type="password", key="cambio_pass_nueva"
            )
            pass_nueva_confirmar = st.text_input(
                "Confirma la nueva contraseña:",
                type="password",
                key="cambio_pass_confirmar",
            )
            if st.button("Actualizar Contraseña", key="btn_cambiar_pass"):
                password_vigente_actual = obtener_password_empleado(
                    supabase,
                    st.session_state.empresa_id,
                    datos_emp["dni"],
                    str(datos_emp["password"]),
                )
                if not clave_coincide(
                    pass_actual.strip(), password_vigente_actual
                ):
                    st.error("La contraseña actual no es correcta.")
                elif not pass_nueva.strip():
                    st.warning("Escribe la nueva contraseña.")
                elif pass_nueva.strip() != pass_nueva_confirmar.strip():
                    st.error("Las dos contraseñas nuevas no coinciden.")
                elif not supabase:
                    st.warning(
                        "No se pudo guardar: el cliente de Supabase no"
                        " está configurado."
                    )
                else:
                    try:
                        guardar_password_empleado(
                            supabase,
                            st.session_state.empresa_id,
                            datos_emp["dni"],
                            _hash_clave(pass_nueva.strip()),
                        )
                        st.success(
                            "✅ Contraseña actualizada. Úsala la próxima vez"
                            " que inicies sesión."
                        )
                    except Exception as e_pass:
                        st.error(f"No se pudo guardar el cambio: {e_pass}")

        st.divider()

        location = get_geolocation()

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("### 1. Identificación y Ubicación")
            hora_oficial, hora_salida_oficial = obtener_horario_oficial(
                datos_emp, df_sedes, hoy
            )

            st.info(
                f"📌 Horario Pactado para hoy ({DIAS_SEMANA_MAP[hoy.weekday()]}):"
                f" **{hora_oficial}** a **{hora_salida_oficial}**"
            )

            tipo_marcacion = st.radio(
                "Acción a registrar:", ["Entrada", "Salida"], horizontal=True
            )

            fecha_hoy_str = hoy.strftime("%Y-%m-%d")
            ya_marcado = ya_marco_hoy(
                supabase,
                st.session_state.empresa_id,
                datos_emp.get("dni", ""),
                datos_emp["nombre"],
                fecha_hoy_str,
                tipo_marcacion,
            )
            if ya_marcado:
                st.warning(
                    f"⚠️ Ya registraste tu **{tipo_marcacion}** de hoy. Solo se"
                    " permite una Entrada y una Salida por día."
                )

            en_rango = False
            sede_detectada = "Desconocida"
            distancia = 0.0

            try:
                sedes_autorizadas = json.loads(
                    datos_emp.get("sedes_autorizadas", "[]")
                )
            except Exception:
                sedes_autorizadas = [datos_emp.get("sede_principal")]

            if not sedes_autorizadas:
                sedes_autorizadas = [datos_emp.get("sede_principal")]

            if location and "coords" in location:
                lat_user = location["coords"]["latitude"]
                lon_user = location["coords"]["longitude"]

                sede_detectada, distancia, _, _, rango_permitido = (
                    validar_ubicacion(
                        lat_user, lon_user, df_sedes, sedes_autorizadas
                    )
                )

                en_rango = distancia <= rango_permitido
                if en_rango:
                    st.success(
                        f"📍 Ubicación confirmada en: **{sede_detectada}**"
                        f" ({round(distancia, 1)}m - En rango)."
                    )
                else:
                    st.error(
                        f"⚠️ Fuera del rango permitido para tus sedes"
                        f" asignadas: **{sede_detectada}** ({round(distancia, 1)}m"
                        f" - Máx: {rango_permitido}m)."
                    )
            else:
                st.warning(
                    "📍 Obteniendo ubicación GPS real del navegador... Por"
                    " favor, permite el acceso a tu ubicación si el navegador"
                    " lo solicita."
                )

        with col2:
            st.markdown("### 2. Foto Obligatoria")
            img_file = st.camera_input(
                "Toma una foto para confirmar tu identidad"
            )

            btn_disabled = not en_rango or img_file is None or ya_marcado

            if not en_rango and location:
                st.error(
                    "🚫 Bloqueado: No estás dentro del rango de ninguna de tus"
                    " sedes autorizadas."
                )

            if st.button(
                "✅ Confirmar Marcación",
                disabled=btn_disabled,
                type="primary",
                use_container_width=True,
            ):
                foto_valida, msg_foto = validar_foto_captura(img_file)
                if not foto_valida:
                    st.error(f"📷 {msg_foto}")
                    st.stop()

                now = ahora_peru().replace(tzinfo=None)
                fecha_str = now.strftime("%Y-%m-%d")
                hora_str = now.strftime("%H:%M:%S")

                marca_tiempo_archivo = now.strftime("%H%M%S%f")
                nombre_foto = (
                    f"{DIR_FOTOS}/{st.session_state.empresa_id}_{fecha_str}_{datos_emp['nombre'].replace(' ', '_')}_{tipo_marcacion}_{marca_tiempo_archivo}.png"
                )
                image = Image.open(img_file)
                image.save(nombre_foto)

                estado = "Puntual"
                minutos_tardanza = 0
                minutos_extra = 0

                if tipo_marcacion == "Entrada":
                    h_oficial = datetime.strptime(
                        hora_oficial, "%H:%M:%S"
                    ).time()
                    tol_aplicar = (
                        st.session_state.tolerancia_minutos
                        if hoy >= st.session_state.fecha_inicio_tolerancia
                        else 0
                    )
                    h_limite_dt = datetime.combine(
                        hoy, h_oficial
                    ) + timedelta(minutes=tol_aplicar)

                    if now > h_limite_dt:
                        estado = "Tardanza"
                        minutos_tardanza = int(
                            (
                                now - datetime.combine(hoy, h_oficial)
                            ).total_seconds()
                            / 60
                        )

                elif tipo_marcacion == "Salida":
                    h_salida_ofic = datetime.strptime(
                        hora_salida_oficial, "%H:%M:%S"
                    ).time()
                    if now.time() > h_salida_ofic:
                        t1 = datetime.combine(datetime.today(), now.time())
                        t2 = datetime.combine(datetime.today(), h_salida_ofic)
                        minutos_extra = int((t1 - t2).total_seconds() / 60)

                # --- ENVÍO COMPLETO A SUPABASE (NUBE EFÍMERA) ---
                sync_nube_ok = False
                with st.spinner("📤 Guardando tu marcación y foto..."):
                    if supabase:
                        try:
                            solo_nombre_foto = os.path.basename(nombre_foto)
                            with open(nombre_foto, "rb") as f:
                                supabase.storage.from_("fotos-asistencia").upload(
                                    solo_nombre_foto,
                                    f,
                                    file_options={"upsert": "true"},
                                )

                            data_nube = {
                                "empresa_id": str(st.session_state.empresa_id),
                                "dni": str(datos_emp.get('dni', '')),
                                "nombre": str(datos_emp.get('nombre', '')),
                                "tipo": str(tipo_marcacion),
                                "fecha": str(fecha_str),
                                "hora_registrada": str(hora_str),
                                "hora_entrada_oficial": str(hora_oficial),
                                "hora_salida_oficial": str(hora_salida_oficial),
                                "estado": str(estado),
                                "minutos_tardanza": int(minutos_tardanza),
                                "horas_extra_min": int(minutos_extra),
                                "sede_detectada": str(sede_detectada),
                                "distancia_m": float(round(distancia, 1)),
                                "en_rango": "SÍ" if en_rango else "NO",
                                "foto_url": solo_nombre_foto,
                                "descargado_master": False,
                                "descargado_cliente": False
                            }
                            supabase.table("marcaciones_efimeras").insert(data_nube).execute()
                            sync_nube_ok = True
                        except Exception as err_nube:
                            msg = str(err_nube)
                            if "Bucket not found" in msg:
                                st.error(
                                    "☁️ No se pudo sincronizar con la Nube: el bucket "
                                    "'fotos-asistencia' no existe en tu proyecto de "
                                    "Supabase. Créalo en Storage → New bucket con ese "
                                    "nombre exacto y agrega políticas de acceso."
                                )
                            elif "row-level security" in msg.lower() or "RLS" in msg or "403" in msg or "401" in msg:
                                st.error(
                                    "☁️ No se pudo sincronizar con la Nube: la petición "
                                    "fue bloqueada por Row Level Security (RLS). Revisa "
                                    "las políticas de la tabla 'marcaciones_efimeras' en "
                                    "Supabase (INSERT/SELECT/UPDATE para el rol usado por "
                                    "tu API key)."
                                )
                            else:
                                st.error(f"☁️ No se pudo sincronizar con la Nube: {msg}")
                    else:
                        st.warning(
                            "☁️ Cliente de Supabase no configurado: la marcación se "
                            "guardará solo localmente."
                        )

                nueva_marcacion = {
                    "empresa_id": st.session_state.empresa_id,
                    "Fecha": fecha_str,
                    "Empleado": datos_emp["nombre"],
                    "Tipo Marcación": tipo_marcacion,
                    "Hora Registrada": hora_str,
                    "Hora Entrada Oficial": hora_oficial,
                    "Hora Salida Oficial": hora_salida_oficial,
                    "Estado": estado,
                    "Minutos Tardanza": minutos_tardanza,
                    "Horas Extra (min)": minutos_extra,
                    "Sede Detectada": sede_detectada,
                    "Distancia (m)": round(distancia, 1),
                    "En Rango": "SÍ" if en_rango else "NO",
                    "Foto": nombre_foto,
                }

                with bloqueo_csv(CSV_ASISTENCIA):
                    archivo_existia = os.path.exists(CSV_ASISTENCIA)
                    df_fila_nueva = pd.DataFrame(
                        [nueva_marcacion]
                    )[COLUMNAS_ASISTENCIA]
                    df_fila_nueva.to_csv(
                        CSV_ASISTENCIA,
                        mode="a" if archivo_existia else "w",
                        header=not archivo_existia,
                        index=False,
                    )

                if sync_nube_ok:
                    st.success(
                        f"¡Marcación de {tipo_marcacion} registrada y "
                        "sincronizada con la Nube!"
                    )
                else:
                    st.success(
                        f"¡Marcación de {tipo_marcacion} registrada "
                        "localmente!"
                    )

                # --- Animación de "globos" con el logo (celebración) ---
                _logo_globos = st.session_state.get(
                    "logo_globos_url", "/app/static/icon-192.png"
                )
                _html_globos = (
                    '<div style="position:fixed; inset:0; pointer-events:none;'
                    ' z-index:9999; overflow:hidden;">'
                )
                for _i in range(10):
                    _left = random.randint(2, 92)
                    _delay = round(random.uniform(0, 1.4), 2)
                    _drift = random.randint(-60, 60)
                    _rot = random.randint(-14, 14)
                    _dur = round(random.uniform(3.4, 5.2), 2)
                    _html_globos += f"""
                    <div style="position:absolute; bottom:-140px; left:{_left}%;
                        width:52px; height:66px;
                        animation:fac-float-up {_dur}s ease-in {_delay}s 1;
                        --drift:{_drift}px; --rot:{_rot}deg;">
                        <div style="width:100%; height:100%;
                            border-radius:50% 50% 50% 50% / 58% 58% 42% 42%;
                            background:linear-gradient(160deg, var(--cyan), var(--violet));
                            box-shadow:0 6px 18px rgba(0,0,0,0.35);
                            display:flex; align-items:center; justify-content:center;">
                            <img src="{_logo_globos}" style="width:60%; height:60%;
                                object-fit:contain; border-radius:50%;
                                background:rgba(255,255,255,0.85);" />
                        </div>
                        <div style="position:absolute; left:50%; top:100%; width:1px;
                            height:24px; background:rgba(255,255,255,0.35);
                            transform:translateX(-50%);"></div>
                    </div>
                    """
                _html_globos += """
                </div>
                <style>
                @keyframes fac-float-up{
                    0%{ transform:translateY(0) translateX(0) rotate(0deg); opacity:0; }
                    8%{ opacity:1; }
                    100%{ transform:translateY(-115vh) translateX(var(--drift, 30px)) rotate(var(--rot, 8deg)); opacity:0; }
                }
                </style>
                """
                st.markdown(_html_globos, unsafe_allow_html=True)

elif opcion == "🔐 Panel de Gestión / Admin":
    if not st.session_state.autenticado:
        st.title("🔐 Acceso Administrativo")

        df_e_disponibles = df_empresas[
            df_empresas["entorno"] == st.session_state.entorno
        ]
        empresa_admin = st.selectbox(
            f"Seleccione Empresa ({st.session_state.entorno}):",
            df_e_disponibles["empresa_id"].unique()
            if not df_e_disponibles.empty
            else [],
        )
        if empresa_admin:
            cargar_configuracion_sistema(supabase, empresa_admin)
        pin = st.text_input("Ingrese PIN de Acceso:", type="password")

        if st.button("Ingresar al Panel"):
            if empresa_admin:
                st.session_state.empresa_id = empresa_admin
                if clave_coincide(pin, st.session_state.pin_admin):
                    st.session_state.autenticado = True
                    st.session_state.rol = "admin"
                    st.rerun()
                elif clave_coincide(pin, st.session_state.pin_visor):
                    st.session_state.autenticado = True
                    st.session_state.rol = "visor"
                    st.rerun()
                elif clave_coincide(pin, st.session_state.pin_master):
                    st.session_state.autenticado = True
                    st.session_state.rol = "master"
                    st.rerun()
                else:
                    st.error("PIN Incorrecto.")
            else:
                st.error("No hay empresas disponibles en este entorno.")
    else:
        from streamlit_autorefresh import st_autorefresh

        # --- Auto-refresh inteligente ---
        # 1) Si la pestaña del navegador no está a la vista (el usuario
        #    cambió a otra ventana/pestaña), se espacía mucho el refresco:
        #    no tiene sentido seguir consultando Supabase si nadie está
        #    mirando la pantalla.
        # 2) Si la pestaña sí está visible pero llevan varios ciclos
        #    seguidos sin ninguna marcación nueva, el intervalo se va
        #    alargando poco a poco (10s → 20s → 40s, tope 60s) para no
        #    machacar Supabase en horas muertas. En cuanto aparece algo
        #    nuevo, vuelve de inmediato a 10s (para no hacer esperar al
        #    admin justo cuando sí está pasando algo).
        if "pestana_visible" not in st.session_state:
            st.session_state.pestana_visible = True
        if "admin_intervalo_autorefresh_ms" not in st.session_state:
            st.session_state.admin_intervalo_autorefresh_ms = 10_000
        if "admin_ciclos_sin_cambios" not in st.session_state:
            st.session_state.admin_ciclos_sin_cambios = 0

        _visible_detectado = streamlit_js_eval(
            js_expressions="document.visibilityState === 'visible'",
            key="PESTANA_VISIBLE",
        )
        if _visible_detectado is not None:
            st.session_state.pestana_visible = bool(_visible_detectado)

        hubo_cambios = st.session_state.pop(
            "_ultima_sync_hubo_cambios", None
        )
        if hubo_cambios is True:
            st.session_state.admin_ciclos_sin_cambios = 0
            st.session_state.admin_intervalo_autorefresh_ms = 10_000
        elif hubo_cambios is False:
            st.session_state.admin_ciclos_sin_cambios += 1
            if st.session_state.admin_ciclos_sin_cambios >= 3:
                st.session_state.admin_intervalo_autorefresh_ms = min(
                    st.session_state.admin_intervalo_autorefresh_ms * 2,
                    60_000,
                )
                st.session_state.admin_ciclos_sin_cambios = 0

        if st.session_state.pestana_visible:
            intervalo_ms = st.session_state.admin_intervalo_autorefresh_ms
        else:
            intervalo_ms = 60_000  # pestaña en segundo plano: casi en pausa

        st_autorefresh(interval=intervalo_ms, key="admin_autorefresh")

        st.title(
            f"⚙️ Control Administrativo - [{st.session_state.empresa_id}]"
            f" ({st.session_state.entorno})"
        )

        _tablas_locales = st.session_state.get("tablas_en_modo_local", set())
        if _tablas_locales:
            _nombres_visibles = {
                "sedes": "Sedes",
                "empleados": "Personal/Trabajadores",
                "empresas": "Empresas",
            }
            _lista_legible = ", ".join(
                _nombres_visibles.get(t, t) for t in sorted(_tablas_locales)
            )
            st.warning(
                f"⚠️ Modo local activo para: **{_lista_legible}**. No se "
                "pudo leer desde la Nube (Supabase) al cargar estos datos; "
                "se está mostrando la última copia guardada en este "
                "servidor. Los cambios recientes hechos desde otro "
                "dispositivo podrían no reflejarse todavía, y si este "
                "servidor se redespliega, los datos aún no sincronizados "
                "se perderían."
            )

        tabs = [
            "🏢 Dashboard General por Local",
            "👤 Reporte Limpio por Trabajador",
        ]

        es_master_o_dev = (
            st.session_state.rol == "master"
            or st.session_state.entorno == "DEV"
        )

        if st.session_state.rol in ["admin", "master"] and not ES_CELULAR:
            if es_master_o_dev:
                tab_gestion_nombre = "🏢 Gestión de Empresas y Sedes"
            else:
                tab_gestion_nombre = "📍 Gestión de Sedes"

            tabs.extend([
                tab_gestion_nombre,
                "👥 Personal",
                "⚙️ Ajustes",
            ])
        elif st.session_state.rol in ["admin", "master"] and ES_CELULAR:
            st.caption(
                "📱 Estás viendo la versión móvil: solo reportes. Entra "
                "desde una laptop/PC para gestionar empresas, personal y"
                " ajustes."
            )

        tab_objs = st.tabs(tabs)

        with tab_objs[0]:
            st.markdown("### 🏢 Resumen Consolidado por Sede (Principal)")

            c_f1, c_f2, c_f3 = st.columns([2, 2, 2])
            with c_f1:
                mes_nombre_sel = st.selectbox(
                    "Mes Evaluado:",
                    list(MESES_NOMBRES.values()),
                    index=ahora_peru().month - 1,
                )
                mes_sel = MESES_INVERSO[mes_nombre_sel]
            with c_f2:
                anio_sel = st.number_input(
                    "Año Evaluado:",
                    min_value=2024,
                    max_value=2030,
                    value=ahora_peru().year,
                )
            with c_f3:
                st.write("")
                st.write("")
                if st.button(
                    "📥 Descargar Reporte Excel Completo",
                    use_container_width=True,
                ):
                    with st.spinner("📊 Generando el Excel completo..."):
                        excel_bytes = generar_excel_completo(
                            df_asistencia,
                            df_empleados,
                            mes_sel,
                            anio_sel,
                            st.session_state.clave_excel,
                        )
                    st.download_button(
                        label="💾 Confirmar Descarga de Excel",
                        data=excel_bytes,
                        file_name=(
                            f"Reporte_Asistencia_{st.session_state.empresa_id}_"
                            f"{mes_nombre_sel}_{anio_sel}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ),
                    )

            st.divider()

            sedes_unicas = (
                list(df_sedes["nombre_sede"].unique())
                if not df_sedes.empty
                else []
            )
            prefix_filtro = f"{anio_sel}-{mes_sel:02d}"
            num_dias_mes_gen = calendar.monthrange(anio_sel, mes_sel)[1]

            for sede in sedes_unicas:
                st.markdown(f"#### 📍 Local Principal: **{sede}**")
                emps_sede = df_empleados[
                    df_empleados["sede_principal"] == sede
                ]

                if emps_sede.empty:
                    st.caption(
                        "No hay personal asignado a esta sede como principal."
                    )
                else:
                    k1, k2, k3, k4 = st.columns(4)
                    tot_p, tot_t, tot_min_t, tot_min_e = 0, 0, 0, 0

                    for _, emp_row in emps_sede.iterrows():
                        df_emp_a = df_asistencia[
                            (df_asistencia["Empleado"] == emp_row["nombre"])
                            & (
                                df_asistencia["Fecha"]
                                .astype(str)
                                .str.startswith(prefix_filtro)
                            )
                        ]
                        if not df_emp_a.empty:
                            tot_p += df_emp_a[df_emp_a["Estado"] == "Puntual"][
                                "Fecha"
                            ].nunique()
                            tot_t += df_emp_a[df_emp_a["Estado"] == "Tardanza"][
                                "Fecha"
                            ].nunique()
                            tot_min_t += df_emp_a["Minutos Tardanza"].sum()
                            tot_min_e += df_emp_a["Horas Extra (min)"].sum()

                    max_dias_posibles = len(emps_sede) * num_dias_mes_gen

                    hrs_dec_sede = round(tot_min_t / 60.0, 1)
                    fmt_hm_sede = min_a_formato_horas(tot_min_t)

                    k1.metric("Personal Afiliado", f"{len(emps_sede)} emps")
                    k2.metric(
                        "Puntualidades Totales",
                        f"{tot_p} días",
                        delta=f"{tot_p} de {max_dias_posibles} días-persona",
                    )
                    
                    if es_mejora_activa(st.session_state.entorno):
                        k3.metric(
                            label="Horas Tardanza Sede",
                            value=f"{hrs_dec_sede} hrs",
                            delta=f"({fmt_hm_sede})",
                            delta_color="off"
                        )
                    else:
                        k3.metric("Horas Tardanza Sede", f"{round(tot_min_t / 60.0, 2)} hrs")

                    k4.metric("Minutos Extras Totales", f"{tot_min_e} min")

                    st.markdown("**Personal de la Sede:**")
                    df_resumen_local = []
                    for _, emp_row in emps_sede.iterrows():
                        df_emp_a = df_asistencia[
                            (df_asistencia["Empleado"] == emp_row["nombre"])
                            & (
                                df_asistencia["Fecha"]
                                .astype(str)
                                .str.startswith(prefix_filtro)
                            )
                        ]
                        p_cnt = (
                            df_emp_a[df_emp_a["Estado"] == "Puntual"][
                                "Fecha"
                            ].nunique()
                            if not df_emp_a.empty
                            else 0
                        )
                        t_cnt = (
                            df_emp_a[df_emp_a["Estado"] == "Tardanza"][
                                "Fecha"
                            ].nunique()
                            if not df_emp_a.empty
                            else 0
                        )
                        m_sum = (
                            df_emp_a["Minutos Tardanza"].sum()
                            if not df_emp_a.empty
                            else 0
                        )
                        e_sum = (
                            df_emp_a["Horas Extra (min)"].sum()
                            if not df_emp_a.empty
                            else 0
                        )

                        df_resumen_local.append({
                            "DNI": emp_row["dni"],
                            "Nombre": emp_row["nombre"],
                            "Cargo": emp_row["cargo"],
                            "Días Puntuales": f"{p_cnt} / {num_dias_mes_gen}",
                            "Tardanzas": t_cnt,
                            "Horas Tardanza": round(m_sum / 60.0, 2),
                            "Min. Extras": e_sum,
                        })
                    st.dataframe(
                        pd.DataFrame(df_resumen_local),
                        use_container_width=True,
                        hide_index=True,
                    )
                st.divider()

        with tab_objs[1]:
            st.markdown("### 👤 Reporte e Inspección Detallada por Trabajador")

            c_e1, c_e2, c_e3 = st.columns([3, 1.5, 1.5])
            with c_e1:
                emp_ind_sel = st.selectbox(
                    "Seleccione el Trabajador:",
                    (
                        df_empleados["nombre"].unique()
                        if not df_empleados.empty
                        else []
                    ),
                )
            with c_e2:
                mes_ind_sel = st.selectbox(
                    "Mes:",
                    list(MESES_NOMBRES.values()),
                    index=ahora_peru().month - 1,
                    key="m_ind_clean",
                )
            with c_e3:
                anio_ind_sel = st.number_input(
                    "Año:",
                    min_value=2024,
                    max_value=2030,
                    value=ahora_peru().year,
                    key="a_ind_clean",
                )

            if emp_ind_sel:
                m_num = MESES_INVERSO[mes_ind_sel]
                prefix_ind = f"{anio_ind_sel}-{m_num:02d}"
                num_dias_m = calendar.monthrange(anio_ind_sel, m_num)[1]

                emp_info = df_empleados[
                    df_empleados["nombre"] == emp_ind_sel
                ].iloc[0]

                try:
                    s_aut_list = json.loads(
                        emp_info.get("sedes_autorizadas", "[]")
                    )
                    s_aut_str = ", ".join(s_aut_list)
                except Exception:
                    s_aut_str = emp_info["sede_principal"]

                st.info(
                    f"📋 **Trabajador:** {emp_info['nombre']} | **DNI:**"
                    f" {emp_info['dni']} | **Cargo:** {emp_info['cargo']} |"
                    f" **Sede Principal:** {emp_info['sede_principal']} |"
                    f" **Sedes Habilitadas:** {s_aut_str}"
                )

                es_autorizado_edicion = (
                    st.session_state.rol in ["admin", "master"]
                    or st.session_state.entorno == "DEV"
                ) and not ES_CELULAR

                if es_autorizado_edicion:
                    with st.expander(
                        "🛠️ Herramientas Avanzadas de Regularización de"
                        " Asistencia (Exclusivo SuperAdmin / Dev)",
                        expanded=False,
                    ):
                        st.caption(
                            "Ideal para trabajadores antiguos que ingresaron"
                            " antes de la implementación del sistema GPS."
                        )

                        col_reg1, col_reg2 = st.columns([1.5, 2])

                        with col_reg1:
                            st.markdown(
                                "##### 1️⃣ Regularización Masiva por Defecto"
                            )
                            st.write(
                                "Genera automáticamente marcaciones de"
                                " **Entrada y Salida Puntual** para todos los"
                                " días laborables transcurridos sin registro del"
                                " mes seleccionado."
                            )

                            if st.button(
                                "⚡ Marcar Entrada/Salida Puntual Masiva",
                                use_container_width=True,
                            ):
                                df_asist_all = (
                                    pd.read_csv(CSV_ASISTENCIA)
                                    if os.path.exists(CSV_ASISTENCIA)
                                    else pd.DataFrame()
                                )
                                hoy_actual = hoy_peru()
                                h_ent_ofic, h_sal_ofic = (
                                    obtener_horario_oficial(
                                        emp_info, df_sedes, hoy_actual
                                    )
                                )

                                nuevos_registros_regularizados = []
                                cant_creados = 0

                                for d_reg in range(1, num_dias_m + 1):
                                    f_reg = date(anio_ind_sel, m_num, d_reg)
                                    f_reg_str = f_reg.strftime("%Y-%m-%d")
                                    nom_d_reg = DIAS_SEMANA_MAP[f_reg.weekday()]

                                    if (
                                        f_reg <= hoy_actual
                                        and f_reg_str not in FERIADOS_OFICIALES
                                        and nom_d_reg
                                        in st.session_state.dias_laborables
                                    ):
                                        existe = (
                                            not df_asist_all.empty
                                            and not df_asist_all[
                                                (
                                                    df_asist_all["empresa_id"].astype(str)
                                                    == str(st.session_state.empresa_id)
                                                )
                                                & (
                                                    df_asist_all["Empleado"]
                                                    == emp_ind_sel
                                                )
                                                & (
                                                    df_asist_all["Fecha"]
                                                    == f_reg_str
                                                )
                                            ].empty
                                        )

                                        if not existe:
                                            nuevos_registros_regularizados.append(
                                                {
                                                    "empresa_id": (
                                                        st.session_state.empresa_id
                                                    ),
                                                    "Fecha": f_reg_str,
                                                    "Empleado": emp_ind_sel,
                                                    "Tipo Marcación": "Entrada",
                                                    "Hora Registrada": (
                                                        h_ent_ofic
                                                    ),
                                                    "Hora Entrada Oficial": (
                                                        h_ent_ofic
                                                    ),
                                                    "Hora Salida Oficial": (
                                                        h_sal_ofic
                                                    ),
                                                    "Estado": "Puntual",
                                                    "Minutos Tardanza": 0,
                                                    "Horas Extra (min)": 0,
                                                    "Sede Detectada": emp_info[
                                                        "sede_principal"
                                                    ],
                                                    "Distancia (m)": 0.0,
                                                    "En Rango": "SÍ",
                                                    "Foto": "",
                                                }
                                            )
                                            nuevos_registros_regularizados.append(
                                                {
                                                    "empresa_id": (
                                                        st.session_state.empresa_id
                                                    ),
                                                    "Fecha": f_reg_str,
                                                    "Empleado": emp_ind_sel,
                                                    "Tipo Marcación": "Salida",
                                                    "Hora Registrada": (
                                                        h_sal_ofic
                                                    ),
                                                    "Hora Entrada Oficial": (
                                                        h_ent_ofic
                                                    ),
                                                    "Hora Salida Oficial": (
                                                        h_sal_ofic
                                                    ),
                                                    "Estado": "Puntual",
                                                    "Minutos Tardanza": 0,
                                                    "Horas Extra (min)": 0,
                                                    "Sede Detectada": emp_info[
                                                        "sede_principal"
                                                    ],
                                                    "Distancia (m)": 0.0,
                                                    "En Rango": "SÍ",
                                                    "Foto": "",
                                                }
                                            )
                                            cant_creados += 1

                                if nuevos_registros_regularizados:
                                    with bloqueo_csv(CSV_ASISTENCIA):
                                        df_asist_all = (
                                            pd.read_csv(CSV_ASISTENCIA)
                                            if os.path.exists(CSV_ASISTENCIA)
                                            else pd.DataFrame()
                                        )
                                        df_asist_all = pd.concat(
                                            [
                                                df_asist_all,
                                                pd.DataFrame(
                                                    nuevos_registros_regularizados
                                                ),
                                            ],
                                            ignore_index=True,
                                        )
                                        df_asist_all.to_csv(
                                            CSV_ASISTENCIA, index=False
                                        )
                                    st.success(
                                        f"¡Se regularizaron {cant_creados} días"
                                        " como Puntual para"
                                        f" {emp_ind_sel}!"
                                    )
                                    st.rerun()
                                else:
                                    st.info(
                                        "No hay días pendientes de"
                                        " regularización para este mes."
                                    )

                        with col_reg2:
                            st.markdown(
                                "##### 2️⃣ Edición Individual (Solo Días Sin"
                                " Foto)"
                            )

                            df_asist_actual = (
                                pd.read_csv(CSV_ASISTENCIA)
                                if os.path.exists(CSV_ASISTENCIA)
                                else pd.DataFrame()
                            )

                            if not df_asist_actual.empty:
                                mask_ed = (
                                    (
                                        df_asist_actual["empresa_id"].astype(str)
                                        == str(st.session_state.empresa_id)
                                    )
                                    & (
                                        df_asist_actual["Empleado"]
                                        == emp_ind_sel
                                    )
                                    & (
                                        df_asist_actual["Fecha"]
                                        .astype(str)
                                        .str.startswith(prefix_ind)
                                    )
                                    & (
                                        df_asist_actual["Foto"].isna()
                                        | (
                                            df_asist_actual["Foto"]
                                            .astype(str)
                                            .str.strip()
                                            == ""
                                        )
                                    )
                                )
                                df_editables = df_asist_actual[mask_ed]

                                if not df_editables.empty:
                                    fechas_disponibles = sorted(
                                        df_editables["Fecha"].unique()
                                    )
                                    f_edit_sel = st.selectbox(
                                        "Seleccione la Fecha sin foto a"
                                        " modificar:",
                                        fechas_disponibles,
                                    )

                                    col_e_m1, col_e_m2, col_e_m3 = st.columns(3)
                                    with col_e_m1:
                                        nuevo_est = st.selectbox(
                                            "Estado:",
                                            ["Puntual", "Tardanza", "Falta"],
                                        )
                                    with col_e_m2:
                                        nuevos_min_t = st.number_input(
                                            "Min. Tardanza:",
                                            min_value=0,
                                            value=0,
                                        )
                                    with col_e_m3:
                                        nuevos_min_e = st.number_input(
                                            "Min. Extra:",
                                            min_value=0,
                                            value=0,
                                        )

                                    if st.button("💾 Guardar Ajuste Manual"):
                                        with bloqueo_csv(CSV_ASISTENCIA):
                                            df_asist_fresco = (
                                                pd.read_csv(CSV_ASISTENCIA)
                                                if os.path.exists(
                                                    CSV_ASISTENCIA
                                                )
                                                else pd.DataFrame()
                                            )
                                            indices = df_asist_fresco[
                                                (
                                                    df_asist_fresco[
                                                        "empresa_id"
                                                    ].astype(str)
                                                    == str(
                                                        st.session_state.empresa_id
                                                    )
                                                )
                                                & (
                                                    df_asist_fresco["Empleado"]
                                                    == emp_ind_sel
                                                )
                                                & (
                                                    df_asist_fresco["Fecha"]
                                                    == f_edit_sel
                                                )
                                            ].index

                                            for idx_mod in indices:
                                                df_asist_fresco.at[
                                                    idx_mod, "Estado"
                                                ] = nuevo_est
                                                df_asist_fresco.at[
                                                    idx_mod,
                                                    "Minutos Tardanza",
                                                ] = nuevos_min_t
                                                df_asist_fresco.at[
                                                    idx_mod,
                                                    "Horas Extra (min)",
                                                ] = nuevos_min_e

                                            df_asist_fresco.to_csv(
                                                CSV_ASISTENCIA, index=False
                                            )
                                        st.success(
                                            f"Registro del día {f_edit_sel}"
                                            " actualizado con éxito."
                                        )
                                        st.rerun()
                                else:
                                    st.caption(
                                        "🔒 No hay registros sin foto para este"
                                        " mes. Las marcaciones reales tomadas"
                                        " con cámara no pueden ser editadas"
                                        " manualmente."
                                    )

                df_asist_emp = df_asistencia[
                    (df_asistencia["Empleado"] == emp_ind_sel)
                    & (
                        df_asistencia["Fecha"]
                        .astype(str)
                        .str.startswith(prefix_ind)
                    )
                ]

                if not df_asist_emp.empty:
                    total_puntual = df_asist_emp[
                        df_asist_emp["Estado"] == "Puntual"
                    ]["Fecha"].nunique()
                    total_tardanza = df_asist_emp[
                        df_asist_emp["Estado"] == "Tardanza"
                    ]["Fecha"].nunique()
                    minutos_tardanza_acumulados = df_asist_emp[
                        "Minutos Tardanza"
                    ].sum()
                    minutos_extra_acumulados = df_asist_emp[
                        "Horas Extra (min)"
                    ].sum()
                else:
                    total_puntual = 0
                    total_tardanza = 0
                    minutos_tardanza_acumulados = 0
                    minutos_extra_acumulados = 0

                horas_tardanza_dec = round(
                    minutos_tardanza_acumulados / 60.0, 2
                )
                formato_hhmm_tardanza = min_a_formato_horas(
                    minutos_tardanza_acumulados
                )

                m1, m2, m3, m4 = st.columns(4)
                m1.metric(
                    "Días Puntuales (Mes)",
                    f"{total_puntual} días",
                    delta=f"{total_puntual} de {num_dias_m} días ({mes_ind_sel})",
                )
                m2.metric(
                    "Días con Tardanza",
                    f"{total_tardanza} días",
                    delta=f"{total_tardanza} de {num_dias_m} días ({mes_ind_sel})",
                    delta_color="inverse",
                )
                m3.metric(
                    "Horas Tardanza Acumuladas",
                    f"{horas_tardanza_dec} hrs",
                    delta=formato_hhmm_tardanza,
                    delta_color="inverse",
                )
                m4.metric(
                    "Minutos Extras Trabajados",
                    f"{minutos_extra_acumulados} min",
                    delta=min_a_formato_horas(minutos_extra_acumulados),
                )

                st.divider()

                st.markdown("#### 📊 Comportamiento Diario de Asistencia")
                timeline_data = []
                for d in range(1, num_dias_m + 1):
                    f_eval = date(anio_ind_sel, m_num, d)
                    f_str = f_eval.strftime("%Y-%m-%d")

                    df_dia_emp = df_asist_emp[df_asist_emp["Fecha"] == f_str]
                    ent_reg = df_dia_emp[
                        df_dia_emp["Tipo Marcación"] == "Entrada"
                    ]

                    if not ent_reg.empty:
                        est = ent_reg.iloc[0]["Estado"]
                        val_y = 2 if est == "Puntual" else 1
                        timeline_data.append({
                            "Día": d,
                            "Estado": est.upper(),
                            "Nivel": val_y,
                        })

                if timeline_data:
                    df_tl = pd.DataFrame(timeline_data)
                    fig_tl = px.bar(
                        df_tl,
                        x="Día",
                        y="Nivel",
                        color="Estado",
                        color_discrete_map={
                            "PUNTUAL": "#2EB67D",
                            "TARDANZA": "#FF8C00",
                        },
                        height=260,
                    )
                    fig_tl.update_layout(
                        yaxis=dict(
                            tickmode="array",
                            tickvals=[1, 2],
                            ticktext=["TARDANZA", "PUNTUAL"],
                        ),
                        xaxis=dict(dtick=1),
                        margin=dict(l=10, r=10, t=10, b=10),
                    )
                    st.plotly_chart(
                        fig_tl,
                        use_container_width=True,
                        key=f"tl_clean_{emp_ind_sel}",
                    )

                st.markdown(
                    "#### 🔍 Bitácora de Marcaciones, Fotos y Verificación GPS"
                )

                lista_registros_completos = []
                hoy_eval = hoy_peru()

                for d in range(1, num_dias_m + 1):
                    f_curr = date(anio_ind_sel, m_num, d)
                    f_curr_str = f_curr.strftime("%Y-%m-%d")
                    nom_dia = DIAS_SEMANA_MAP[f_curr.weekday()]
                    fecha_formateada = (
                        f"{nom_dia} {d:02d}/{m_num:02d}/{anio_ind_sel}"
                    )

                    if f_curr <= hoy_eval:
                        df_marcas_dia = df_asist_emp[
                            df_asist_emp["Fecha"] == f_curr_str
                        ]

                        if not df_marcas_dia.empty:
                            for _, r_m in df_marcas_dia.iterrows():
                                lista_registros_completos.append({
                                    "Fecha_Raw": f_curr_str,
                                    "Fecha_Display": fecha_formateada,
                                    "Tipo Marcación": r_m["Tipo Marcación"],
                                    "Hora Registrada": r_m["Hora Registrada"],
                                    "Estado": r_m["Estado"],
                                    "Minutos Tardanza": r_m["Minutos Tardanza"],
                                    "Horas Extra (min)": r_m.get(
                                        "Horas Extra (min)", 0
                                    ),
                                    "Sede Detectada": r_m["Sede Detectada"],
                                    "Distancia (m)": r_m["Distancia (m)"],
                                    "En Rango": r_m["En Rango"],
                                    "Foto": r_m.get("Foto", None),
                                    "Es_Especial": False,
                                })
                        else:
                            if f_curr_str in FERIADOS_OFICIALES:
                                lista_registros_completos.append({
                                    "Fecha_Raw": f_curr_str,
                                    "Fecha_Display": fecha_formateada,
                                    "Tipo Marcación": "—",
                                    "Hora Registrada": "—",
                                    "Estado": (
                                        f"FERIADO ({FERIADOS_OFICIALES[f_curr_str]})"
                                    ),
                                    "Minutos Tardanza": 0,
                                    "Horas Extra (min)": 0,
                                    "Sede Detectada": "—",
                                    "Distancia (m)": "—",
                                    "En Rango": "—",
                                    "Foto": None,
                                    "Es_Especial": True,
                                })
                            elif nom_dia in st.session_state.dias_laborables:
                                lista_registros_completos.append({
                                    "Fecha_Raw": f_curr_str,
                                    "Fecha_Display": fecha_formateada,
                                    "Tipo Marcación": "Entrada / Salida",
                                    "Hora Registrada": "Sin Registro",
                                    "Estado": "FALTA",
                                    "Minutos Tardanza": 0,
                                    "Horas Extra (min)": 0,
                                    "Sede Detectada": "—",
                                    "Distancia (m)": "—",
                                    "En Rango": "—",
                                    "Foto": None,
                                    "Es_Especial": True,
                                })

                if lista_registros_completos:
                    st.markdown(
                        render_custom_table(lista_registros_completos),
                        unsafe_allow_html=True,
                    )
                else:
                    st.warning(
                        "No existen registros o evaluaciones disponibles para"
                        " este mes."
                    )

        if st.session_state.rol in ["admin", "master"] and not ES_CELULAR:
            with tab_objs[2]:
                if es_master_o_dev:
                    st.subheader("🏢 Gestión Integral de Empresas y Sedes SaaS")
                    tab_sub_sedes, tab_sub_empresas = st.tabs([
                        "📍 Sedes de la Empresa Activa",
                        "🏢 Gestión de Empresas (Dev / SaaS)",
                    ])

                    with tab_sub_sedes:
                        render_modulo_sedes(df_sedes)

                    with tab_sub_empresas:
                        render_modulo_empresas()
                else:
                    st.subheader(
                        f"📍 Gestión de Sedes - {st.session_state.empresa_id}"
                    )
                    render_modulo_sedes(df_sedes)

            with tab_objs[3]:
                st.subheader("👥 Gestión de Personal")

                c_p1, c_p2 = st.columns([2.5, 1.5])
                with c_p1:
                    df_emp_display = df_empleados.copy()
                    if "sedes_autorizadas" in df_emp_display.columns:
                        df_emp_display["sedes_autorizadas"] = (
                            df_emp_display["sedes_autorizadas"].apply(
                                lambda x: (
                                    ", ".join(json.loads(x))
                                    if isinstance(x, str)
                                    and x.startswith("[")
                                    else str(x)
                                )
                            )
                        )
                    st.dataframe(
                        df_emp_display[[
                            "dni",
                            "nombre",
                            "cargo",
                            "sede_principal",
                            "sedes_autorizadas",
                            "fecha_ingreso",
                        ]],
                        use_container_width=True,
                        hide_index=True,
                    )

                with c_p2:
                    st.markdown("#### 📝 Crear / Editar Trabajador")

                    opciones_emp = ["-- Registrar Nuevo Empleado --"] + [
                        f"{row['dni']} - {row['nombre']}"
                        for _, row in df_empleados.iterrows()
                    ]
                    emp_sel_ed = st.selectbox(
                        "Seleccionar para Editar:", opciones_emp
                    )

                    is_edit_e = emp_sel_ed != "-- Registrar Nuevo Empleado --"

                    sedes_lista = (
                        list(df_sedes["nombre_sede"].unique())
                        if not df_sedes.empty
                        else []
                    )

                    if is_edit_e:
                        dni_extraido = emp_sel_ed.split(" - ")[0]
                        datos_e = df_empleados[
                            df_empleados["dni"].astype(str) == dni_extraido
                        ].iloc[0]

                        val_dni = str(datos_e["dni"])
                        val_nom = str(datos_e["nombre"])
                        val_car = str(datos_e["cargo"])
                        val_sed_p = (
                            str(datos_e.get("sede_principal", ""))
                            if pd.notna(datos_e.get("sede_principal"))
                            else ""
                        )

                        try:
                            raw_sedes = json.loads(
                                datos_e.get("sedes_autorizadas", "[]")
                            )
                            if isinstance(raw_sedes, list):
                                val_sed_a = [
                                    str(s)
                                    for s in raw_sedes
                                    if pd.notna(s)
                                    and str(s).strip() != ""
                                    and str(s) != "None"
                                ]
                            else:
                                val_sed_a = [val_sed_p] if val_sed_p else []
                        except Exception:
                            val_sed_a = [val_sed_p] if val_sed_p else []

                        val_pas = ""  # nunca se muestra el hash guardado
                    else:
                        val_dni = ""
                        val_nom = ""
                        val_car = ""
                        val_sed_p = sedes_lista[0] if sedes_lista else ""
                        val_sed_a = sedes_lista.copy()
                        val_pas = PASSWORD_EMPLEADO_DEFAULT

                    default_sedes_validas = [
                        s for s in val_sed_a if s in sedes_lista
                    ]

                    e_dni = st.text_input(
                        "DNI:", value=val_dni, disabled=is_edit_e
                    )
                    e_nombre = st.text_input("Nombre Completo:", value=val_nom)
                    e_cargo = st.text_input("Cargo:", value=val_car)

                    idx_sede_default = (
                        sedes_lista.index(val_sed_p)
                        if val_sed_p in sedes_lista
                        else 0
                    )
                    e_sede_principal = st.selectbox(
                        "Sede Principal (Para Reportes):",
                        sedes_lista,
                        index=idx_sede_default if sedes_lista else 0,
                    )

                    e_sedes_autorizadas = st.multiselect(
                        "Sedes Autorizadas para Marcar:",
                        options=sedes_lista,
                        default=default_sedes_validas,
                    )

                    e_pass = st.text_input(
                        "Contraseña Marcación:"
                        + (
                            " (déjala en blanco para no cambiarla)"
                            if is_edit_e
                            else ""
                        ),
                        value=val_pas,
                        type="password",
                    )

                    col_btn_e1, col_btn_e2 = st.columns(2)

                    with col_btn_e1:
                        if is_edit_e:
                            if st.button(
                                "💾 Actualizar Datos",
                                use_container_width=True,
                            ):
                                sedes_finales = (
                                    e_sedes_autorizadas
                                    if e_sedes_autorizadas
                                    else [e_sede_principal]
                                )
                                if e_sede_principal not in sedes_finales:
                                    sedes_finales.append(e_sede_principal)

                                datos_actualizados = {
                                    "empresa_id": st.session_state.empresa_id,
                                    "dni": val_dni,
                                    "nombre": e_nombre.strip().upper(),
                                    "cargo": e_cargo.strip().upper(),
                                    "sede_principal": e_sede_principal,
                                    "sedes_autorizadas": json.dumps(
                                        sedes_finales
                                    ),
                                }
                                if e_pass.strip():
                                    # Solo se toca la contraseña si el
                                    # admin escribió una nueva; si la dejó
                                    # en blanco, la que ya estaba guardada
                                    # no se modifica.
                                    datos_actualizados["password"] = (
                                        _hash_clave(e_pass.strip())
                                    )

                                if supabase:
                                    try:
                                        guardar_empleado_supabase(
                                            supabase, datos_actualizados
                                        )
                                    except Exception as e:
                                        st.warning(
                                            "No se pudo guardar en la nube"
                                            f" ({e}). Se guardó solo local;"
                                            " se perderá en el próximo"
                                            " redespliegue."
                                        )

                                # Se actualiza también el CSV local (caché
                                # inmediata / respaldo offline).
                                if os.path.exists(CSV_EMPLEADOS):
                                    with bloqueo_csv(CSV_EMPLEADOS):
                                        df_emp_full = pd.read_csv(CSV_EMPLEADOS)
                                        df_emp_full["dni"] = df_emp_full[
                                            "dni"
                                        ].astype(str)
                                        idx_e = df_emp_full[
                                            (
                                                df_emp_full[
                                                    "empresa_id"
                                                ].astype(str)
                                                == str(
                                                    st.session_state.empresa_id
                                                )
                                            )
                                            & (
                                                df_emp_full["dni"].astype(str)
                                                == val_dni
                                            )
                                        ].index
                                        if len(idx_e) > 0:
                                            for campo, valor in (
                                                datos_actualizados.items()
                                            ):
                                                df_emp_full.at[
                                                    idx_e[0], campo
                                                ] = valor
                                            df_emp_full.to_csv(
                                                CSV_EMPLEADOS, index=False
                                            )

                                st.success(
                                    "Datos del trabajador actualizados."
                                )
                                st.rerun()
                        else:
                            if st.button(
                                "➕ Agregar Empleado",
                                use_container_width=True,
                            ):
                                if e_dni and e_nombre:
                                    if (
                                        (
                                            df_empleados["empresa_id"].astype(
                                                str
                                            )
                                            == str(
                                                st.session_state.empresa_id
                                            )
                                        )
                                        & (
                                            df_empleados["dni"].astype(str)
                                            == e_dni.strip()
                                        )
                                    ).any():
                                        st.error(
                                            "El DNI ingresado ya existe en esta"
                                            " empresa."
                                        )
                                    else:
                                        sedes_finales = (
                                            e_sedes_autorizadas
                                            if e_sedes_autorizadas
                                            else [e_sede_principal]
                                        )
                                        if (
                                            e_sede_principal
                                            not in sedes_finales
                                        ):
                                            sedes_finales.append(
                                                e_sede_principal
                                            )

                                        nuevo_emp = {
                                            "empresa_id": (
                                                st.session_state.empresa_id
                                            ),
                                            "dni": e_dni.strip(),
                                            "nombre": e_nombre.strip().upper(),
                                            "sede_principal": e_sede_principal,
                                            "sedes_autorizadas": json.dumps(
                                                sedes_finales
                                            ),
                                            "cargo": e_cargo.strip().upper(),
                                            "password": _hash_clave(
                                                e_pass.strip()
                                                if e_pass.strip()
                                                else PASSWORD_EMPLEADO_DEFAULT
                                            ),
                                            "horario_personalizado": "{}",
                                            "fecha_ingreso": (
                                                hoy_peru().strftime(
                                                    "%Y-%m-%d"
                                                )
                                            ),
                                        }

                                        if supabase:
                                            try:
                                                guardar_empleado_supabase(
                                                    supabase, nuevo_emp
                                                )
                                            except Exception as e:
                                                st.warning(
                                                    "No se pudo guardar en"
                                                    f" la nube ({e}). Se"
                                                    " guardó solo local; se"
                                                    " perderá en el próximo"
                                                    " redespliegue."
                                                )

                                        # Copia local (caché / respaldo
                                        # offline).
                                        with bloqueo_csv(CSV_EMPLEADOS):
                                            if os.path.exists(CSV_EMPLEADOS):
                                                df_emp_full = pd.read_csv(
                                                    CSV_EMPLEADOS
                                                )
                                            else:
                                                df_emp_full = pd.DataFrame()
                                            df_emp_full = pd.concat(
                                                [
                                                    df_emp_full,
                                                    pd.DataFrame([nuevo_emp]),
                                                ],
                                                ignore_index=True,
                                            )
                                            df_emp_full.to_csv(
                                                CSV_EMPLEADOS, index=False
                                            )
                                        st.success(
                                            "Empleado registrado"
                                            " correctamente."
                                        )
                                        st.rerun()

                    with col_btn_e2:
                        if is_edit_e:
                            if st.button(
                                "🗑️ Eliminar Trabajador",
                                use_container_width=True,
                            ):
                                if supabase:
                                    try:
                                        eliminar_empleado_supabase(
                                            supabase,
                                            st.session_state.empresa_id,
                                            val_dni,
                                        )
                                    except Exception as e:
                                        st.warning(
                                            "No se pudo eliminar en la nube"
                                            f" ({e}). Se eliminó solo local."
                                        )

                                if os.path.exists(CSV_EMPLEADOS):
                                    with bloqueo_csv(CSV_EMPLEADOS):
                                        df_emp_full = pd.read_csv(CSV_EMPLEADOS)
                                        df_emp_full = df_emp_full[
                                            ~(
                                                (
                                                    df_emp_full[
                                                        "empresa_id"
                                                    ].astype(str)
                                                    == str(
                                                        st.session_state.empresa_id
                                                    )
                                                )
                                                & (
                                                    df_emp_full["dni"].astype(str)
                                                    == val_dni
                                                )
                                            )
                                        ]
                                        df_emp_full.to_csv(
                                            CSV_EMPLEADOS, index=False
                                        )
                                st.warning("Trabajador eliminado.")
                                st.rerun()

                st.divider()
                st.markdown("#### ⚙️ Personalizar Horario por Trabajador")
                emp_h_sel = st.selectbox(
                    "Seleccionar Empleado:", df_empleados["nombre"].unique()
                )

                if emp_h_sel:
                    emp_h_idx = df_empleados[
                        df_empleados["nombre"] == emp_h_sel
                    ].index[0]
                    emp_h_row = df_empleados.loc[emp_h_idx]

                    try:
                        h_dict_actual = json.loads(
                            emp_h_row.get("horario_personalizado", "{}")
                        )
                    except Exception:
                        h_dict_actual = {}

                    st.caption(
                        "Define horarios específicos por día para este"
                        " trabajador (Sobrescribe el horario de la sede)."
                    )

                    nuevo_h_dict = {}
                    cols_dias = st.columns(6)
                    dias_semana = [
                        "Lunes",
                        "Martes",
                        "Miércoles",
                        "Jueves",
                        "Viernes",
                        "Sábado",
                    ]

                    for idx_d, dia in enumerate(dias_semana):
                        with cols_dias[idx_d]:
                            st.markdown(f"**{dia}**")
                            activo = st.checkbox(
                                "Aplica",
                                value=h_dict_actual.get(dia, {}).get(
                                    "activo", True
                                ),
                                key=f"chk_{dia}",
                            )
                            val_ent = h_dict_actual.get(dia, {}).get(
                                "entrada", "08:00:00"
                            )
                            val_sal = h_dict_actual.get(dia, {}).get(
                                "salida", "17:00:00"
                            )

                            t_ent = st.time_input(
                                "Entrada",
                                value=datetime.strptime(
                                    val_ent, "%H:%M:%S"
                                ).time(),
                                key=f"ent_{dia}",
                            )
                            t_sal = st.time_input(
                                "Salida",
                                value=datetime.strptime(
                                    val_sal, "%H:%M:%S"
                                ).time(),
                                key=f"sal_{dia}",
                            )

                            nuevo_h_dict[dia] = {
                                "activo": activo,
                                "entrada": t_ent.strftime("%H:%M:%S"),
                                "salida": t_sal.strftime("%H:%M:%S"),
                            }

                    if st.button("💾 Guardar Horario Personalizado"):
                        dni_h = str(emp_h_row["dni"])
                        horario_json = json.dumps(nuevo_h_dict)

                        if supabase:
                            try:
                                guardar_empleado_supabase(
                                    supabase,
                                    {
                                        "empresa_id": (
                                            st.session_state.empresa_id
                                        ),
                                        "dni": dni_h,
                                        "horario_personalizado": (
                                            horario_json
                                        ),
                                    },
                                )
                            except Exception as e:
                                st.warning(
                                    "No se pudo guardar en la nube"
                                    f" ({e}). Se guardó solo local; se"
                                    " perderá en el próximo redespliegue."
                                )

                        if os.path.exists(CSV_EMPLEADOS):
                            with bloqueo_csv(CSV_EMPLEADOS):
                                df_emp_full = pd.read_csv(CSV_EMPLEADOS)
                                df_emp_full["dni"] = df_emp_full["dni"].astype(
                                    str
                                )
                                idx_h = df_emp_full[
                                    (
                                        df_emp_full["empresa_id"].astype(str)
                                        == str(st.session_state.empresa_id)
                                    )
                                    & (df_emp_full["dni"] == dni_h)
                                ].index
                                if len(idx_h) > 0:
                                    df_emp_full.at[
                                        idx_h[0], "horario_personalizado"
                                    ] = horario_json
                                    df_emp_full.to_csv(
                                        CSV_EMPLEADOS, index=False
                                    )
                        st.success(
                            "Horario individual guardado correctamente."
                        )

            with tab_objs[4]:
                st.subheader(
                    "⚙️ Configuración, Reglas de Asistencia y Respaldo Maestro"
                )

                subtab_reglas, subtab_respaldo, subtab_seguridad = st.tabs([
                    "📅 Reglas y Tolerancia",
                    "☁️ Sincronización y Respaldo Maestro",
                    "🔑 Seguridad y Claves",
                ])

                with subtab_reglas:
                    c_r1, c_r2 = st.columns(2)
                    with c_r1:
                        st.markdown("#### Tolerancia de Entrada")
                        tol_ingresada = st.number_input(
                            "Minutos de Gracia:",
                            value=st.session_state.tolerancia_minutos,
                            min_value=0,
                            max_value=60,
                        )
                        f_ini_tol = st.date_input(
                            "Vigente a partir de:",
                            value=st.session_state.fecha_inicio_tolerancia,
                        )

                        if st.button("Actualizar Regla de Tolerancia"):
                            st.session_state.tolerancia_minutos = tol_ingresada
                            st.session_state.fecha_inicio_tolerancia = (
                                f_ini_tol
                            )
                            st.success("Regla de tolerancia actualizada.")

                    with c_r2:
                        st.markdown("#### Reglas Complementarias")
                        st.info(
                            "Todas las tardanzas se convierten a formato decimal"
                            " y horas para consolidar informes."
                        )

                with subtab_respaldo:
                    st.markdown("#### ☁️ Panel Maestro de SuperAdmin / Respaldo")
                    st.caption(
                        "Sincronización, Respaldo Histórico y Vaciado de"
                        " Nube Efímera."
                    )

                    st.info(
                        "📋 **Cómo funciona este respaldo:** al generarlo, "
                        "se prepara un archivo .zip (Excel + fotos) para "
                        "que lo descargues a esta computadora, **y se envía "
                        "automáticamente una copia por correo al equipo que "
                        "da soporte y mantenimiento a este sistema**, como "
                        "parte del servicio. Recién después de que confirmes "
                        "que ya descargaste el archivo, se eliminan esos "
                        "registros de la nube para mantenerte dentro del "
                        "plan gratuito de Supabase."
                    )

                    st.markdown(
                        "**Empresa activa:** "
                        f"`{st.session_state.empresa_id}`"
                    )

                    if "backup_listo" not in st.session_state:
                        st.session_state.backup_listo = False
                        st.session_state.backup_zip_bytes = None
                        st.session_state.backup_ids_a_borrar = []
                        st.session_state.backup_fotos_a_borrar = []
                        st.session_state.backup_email_ok = False

                    if st.button(
                        "1️⃣ Generar Respaldo (local + copia por correo)",
                        use_container_width=True,
                        disabled=st.session_state.backup_listo,
                    ):
                        if not supabase:
                            st.warning(
                                "El cliente Supabase no está configurado."
                            )
                        else:
                            try:
                                _spinner_backup = st.empty()
                                _spinner_backup.info(
                                    "⏳ Generando el respaldo (Excel +"
                                    " fotos) y enviando la copia por"
                                    " correo... esto puede tardar unos"
                                    " segundos, no cierres esta pestaña."
                                )
                                res = (
                                    supabase.table("marcaciones_efimeras")
                                    .select("*")
                                    .eq(
                                        "empresa_id",
                                        str(st.session_state.empresa_id),
                                    )
                                    .eq("descargado_master", False)
                                    .execute()
                                )
                                registros = res.data or []

                                if not registros:
                                    st.info(
                                        "No hay marcaciones nuevas"
                                        " pendientes de respaldo."
                                    )
                                else:
                                    df_export = pd.DataFrame([
                                        {
                                            "Fecha": r.get("fecha"),
                                            "Empleado": r.get("nombre"),
                                            "DNI": r.get("dni"),
                                            "Tipo": r.get("tipo"),
                                            "Hora Registrada": r.get(
                                                "hora_registrada"
                                            ),
                                            "Estado": r.get("estado"),
                                            "Minutos Tardanza": r.get(
                                                "minutos_tardanza", 0
                                            ),
                                            "Horas Extra (min)": r.get(
                                                "horas_extra_min", 0
                                            ),
                                            "Sede Detectada": r.get(
                                                "sede_detectada"
                                            ),
                                            "Distancia (m)": r.get(
                                                "distancia_m", 0.0
                                            ),
                                            "En Rango": r.get("en_rango"),
                                            "Foto": r.get("foto_url"),
                                        }
                                        for r in registros
                                    ])

                                    buffer_excel = io.BytesIO()
                                    df_export.to_excel(
                                        buffer_excel,
                                        index=False,
                                        engine="openpyxl",
                                    )
                                    buffer_excel.seek(0)

                                    buffer_zip = io.BytesIO()
                                    fotos_ok = []
                                    with zipfile.ZipFile(
                                        buffer_zip,
                                        "w",
                                        zipfile.ZIP_DEFLATED,
                                    ) as zf:
                                        zf.writestr(
                                            f"asistencia_{st.session_state.empresa_id}.xlsx",
                                            buffer_excel.getvalue(),
                                        )
                                        for r in registros:
                                            foto_nombre = r.get("foto_url")
                                            if not foto_nombre:
                                                continue
                                            try:
                                                data_foto = supabase.storage.from_(
                                                    "fotos-asistencia"
                                                ).download(foto_nombre)
                                                zf.writestr(
                                                    f"fotos/{foto_nombre}",
                                                    data_foto,
                                                )
                                                fotos_ok.append(foto_nombre)
                                            except Exception as err_foto:
                                                st.caption(
                                                    "⚠️ No se pudo incluir"
                                                    f" la foto {foto_nombre}:"
                                                    f" {err_foto}"
                                                )

                                    buffer_zip.seek(0)
                                    zip_bytes = buffer_zip.getvalue()

                                    try:
                                        enviar_backup_email(
                                            asunto=(
                                                "Respaldo asistencia - "
                                                f"{st.session_state.empresa_id}"
                                                f" - {hoy_peru()}"
                                            ),
                                            cuerpo=(
                                                f"Respaldo automático de"
                                                f" {len(registros)} registros"
                                                " de la empresa"
                                                f" {st.session_state.empresa_id}."
                                            ),
                                            adjuntos=[(
                                                f"respaldo_{st.session_state.empresa_id}_{hoy_peru()}.zip",
                                                zip_bytes,
                                                "application/zip",
                                            )],
                                        )
                                        email_ok = True
                                    except Exception as err_email:
                                        email_ok = False
                                        st.error(
                                            "⚠️ No se pudo enviar la copia"
                                            f" por correo: {err_email}"
                                        )

                                    st.session_state.backup_listo = True
                                    st.session_state.backup_zip_bytes = (
                                        zip_bytes
                                    )
                                    st.session_state.backup_ids_a_borrar = [
                                        r["id"] for r in registros
                                    ]
                                    st.session_state.backup_fotos_a_borrar = (
                                        fotos_ok
                                    )
                                    st.session_state.backup_email_ok = (
                                        email_ok
                                    )
                                    _spinner_backup.empty()
                                    st.success(
                                        "✅ Respaldo generado con"
                                        f" {len(registros)} registros."
                                        + (
                                            " Copia enviada por correo."
                                            if email_ok
                                            else " (la copia por correo"
                                            " falló, revisa el aviso"
                                            " de arriba)"
                                        )
                                    )
                            except Exception as e_sync:
                                _spinner_backup.empty()
                                st.error(
                                    "Error al generar el respaldo:"
                                    f" {e_sync}"
                                )

                    if (
                        st.session_state.backup_listo
                        and st.session_state.backup_zip_bytes
                    ):
                        st.download_button(
                            "⬇️ Descargar respaldo a esta computadora",
                            data=st.session_state.backup_zip_bytes,
                            file_name=(
                                f"respaldo_{st.session_state.empresa_id}"
                                f"_{hoy_peru()}.zip"
                            ),
                            mime="application/zip",
                            use_container_width=True,
                        )

                        st.warning(
                            "⚠️ Antes de continuar, asegúrate de haber"
                            " descargado y guardado el archivo .zip en un"
                            " lugar seguro. El siguiente paso es"
                            " IRREVERSIBLE: borra estos registros de la"
                            " nube."
                        )

                        if not st.session_state.backup_email_ok:
                            st.error(
                                "🔒 La copia por correo de este respaldo"
                                " falló (revisa el aviso más arriba). Por"
                                " seguridad, no se puede limpiar la nube"
                                " hasta tener al menos una copia"
                                " garantizada fuera de Supabase. Genera el"
                                " respaldo de nuevo, o revisa la"
                                " configuración de correo (EMAIL_REMITENTE"
                                " / EMAIL_APP_PASSWORD en Secrets)."
                            )

                        confirmar_descarga = st.checkbox(
                            "Confirmo que ya descargué y guardé el .zip en"
                            " un lugar seguro."
                        )

                        if st.button(
                            "2️⃣ Ya descargué el respaldo — Limpiar la Nube",
                            use_container_width=True,
                            type="primary",
                            disabled=not (
                                st.session_state.backup_email_ok
                                and confirmar_descarga
                            ),
                        ):
                            try:
                                ids = st.session_state.backup_ids_a_borrar
                                fotos = (
                                    st.session_state.backup_fotos_a_borrar
                                )

                                if fotos:
                                    supabase.storage.from_(
                                        "fotos-asistencia"
                                    ).remove(fotos)

                                for rid in ids:
                                    supabase.table(
                                        "marcaciones_efimeras"
                                    ).delete().eq("id", rid).execute()

                                st.session_state.backup_listo = False
                                st.session_state.backup_zip_bytes = None
                                st.session_state.backup_ids_a_borrar = []
                                st.session_state.backup_fotos_a_borrar = []

                                st.success(
                                    f"🧹 Nube limpiada: {len(ids)} registros"
                                    f" y {len(fotos)} fotos eliminados."
                                )
                                st.rerun()
                            except Exception as e_purge:
                                st.error(
                                    f"Error al limpiar la nube: {e_purge}"
                                )

                    st.divider()
                    st.markdown(
                        "#### 📂 Visor de Histórico (respaldos ya"
                        " descargados)"
                    )
                    st.caption(
                        "Sube aquí un .zip de respaldo que ya hayas"
                        " descargado antes con el botón de arriba. Se abre"
                        " solo en tu navegador, sin tocar Supabase — así"
                        " puedes seguir viendo, filtrando por mes y"
                        " descargando esos períodos aunque ya se hayan"
                        " borrado de la nube."
                    )

                    archivo_historico = st.file_uploader(
                        "Subir archivo de respaldo (.zip):",
                        type=["zip"],
                        key="uploader_historico",
                    )

                    if archivo_historico is not None:
                        try:
                            with zipfile.ZipFile(archivo_historico) as zf_h:
                                nombres_excel = [
                                    n
                                    for n in zf_h.namelist()
                                    if n.lower().endswith(".xlsx")
                                ]
                                if not nombres_excel:
                                    st.error(
                                        "Este .zip no contiene un Excel de"
                                        " asistencia reconocible."
                                    )
                                else:
                                    with zf_h.open(nombres_excel[0]) as f_x:
                                        df_historico = pd.read_excel(f_x)

                                    st.success(
                                        f"✅ Se cargaron"
                                        f" {len(df_historico)} registros"
                                        " del histórico."
                                    )

                                    df_mostrar_h = df_historico
                                    if "Fecha" in df_historico.columns:
                                        df_historico["Fecha"] = (
                                            pd.to_datetime(
                                                df_historico["Fecha"],
                                                errors="coerce",
                                            )
                                        )
                                        meses_disp_h = sorted(
                                            df_historico["Fecha"]
                                            .dt.to_period("M")
                                            .astype(str)
                                            .dropna()
                                            .unique()
                                        )
                                        mes_hist_sel = st.selectbox(
                                            "Filtrar por mes:",
                                            ["Todos"] + list(meses_disp_h),
                                            key="mes_hist_sel",
                                        )
                                        if mes_hist_sel != "Todos":
                                            df_mostrar_h = df_historico[
                                                df_historico["Fecha"]
                                                .dt.to_period("M")
                                                .astype(str)
                                                == mes_hist_sel
                                            ]

                                    st.dataframe(
                                        df_mostrar_h,
                                        use_container_width=True,
                                        hide_index=True,
                                    )

                                    buffer_hist = io.BytesIO()
                                    df_mostrar_h.to_excel(
                                        buffer_hist,
                                        index=False,
                                        engine="openpyxl",
                                    )
                                    buffer_hist.seek(0)
                                    st.download_button(
                                        "⬇️ Descargar este período"
                                        " filtrado (Excel)",
                                        data=buffer_hist.getvalue(),
                                        file_name=(
                                            "historico_filtrado_"
                                            f"{st.session_state.get('mes_hist_sel', 'completo')}.xlsx"
                                        ),
                                        mime=(
                                            "application/vnd.openxmlformats"
                                            "-officedocument.spreadsheetml"
                                            ".sheet"
                                        ),
                                        use_container_width=True,
                                        key="descargar_historico_filtrado",
                                    )

                                    nombres_fotos_h = [
                                        n
                                        for n in zf_h.namelist()
                                        if n.startswith("fotos/")
                                    ]
                                    if nombres_fotos_h:
                                        with st.expander(
                                            "📸 Ver fotos incluidas en este"
                                            f" respaldo ({len(nombres_fotos_h)})"
                                        ):
                                            foto_sel_hist = st.selectbox(
                                                "Elegir foto:",
                                                nombres_fotos_h,
                                                key="foto_sel_hist",
                                            )
                                            if foto_sel_hist:
                                                with zf_h.open(
                                                    foto_sel_hist
                                                ) as f_foto:
                                                    st.image(
                                                        f_foto.read(),
                                                        width=300,
                                                    )
                        except zipfile.BadZipFile:
                            st.error(
                                "El archivo subido no es un .zip válido."
                            )
                        except Exception as e_hist:
                            st.error(
                                f"No se pudo leer el respaldo: {e_hist}"
                            )

                with subtab_seguridad:
                    st.markdown("#### 🔑 Administración de Claves de Acceso")
                    st.caption(
                        "Estos cambios se guardan en Supabase y aplican a"
                        f" la empresa `{st.session_state.empresa_id}`."
                    )
                    st.info(
                        "🔒 Por seguridad, los PINs y contraseñas ya no se"
                        " muestran en pantalla ni se guardan como texto"
                        " plano (se guarda un hash). Deja un campo en"
                        " blanco si no quieres cambiar esa clave."
                    )

                    p_admin = st.text_input(
                        "Nuevo PIN SuperAdmin (déjalo en blanco para no"
                        " cambiarlo):",
                        value="",
                        type="password",
                    )
                    p_visor = st.text_input(
                        "Nuevo PIN Admin (déjalo en blanco para no"
                        " cambiarlo):",
                        value="",
                        type="password",
                    )

                    if st.session_state.rol == "master":
                        p_master = st.text_input(
                            "Nuevo PIN Developer (déjalo en blanco para no"
                            " cambiarlo):",
                            value="",
                            type="password",
                        )
                    else:
                        p_master = ""
                        st.caption(
                            "🔒 El PIN Developer solo es visible y editable"
                            " para quien ingresa con esa clave."
                        )

                    c_excel = st.text_input(
                        "Nueva Clave de Protección Excel (déjala en blanco"
                        " para no cambiarla):",
                        value="",
                        type="password",
                    )

                    if st.button("Guardar Nuevas Claves"):
                        campos_a_guardar = {}
                        if p_admin:
                            campos_a_guardar["pin_admin"] = _hash_clave(
                                p_admin
                            )
                        if p_visor:
                            campos_a_guardar["pin_visor"] = _hash_clave(
                                p_visor
                            )
                        if p_master:
                            campos_a_guardar["pin_master"] = _hash_clave(
                                p_master
                            )
                        if c_excel:
                            # La clave de Excel se guarda tal cual (texto
                            # plano) porque openpyxl necesita el valor
                            # real para proteger la hoja de cálculo; no
                            # es una clave de login como las demás.
                            campos_a_guardar["clave_excel"] = c_excel

                        if not campos_a_guardar:
                            st.info(
                                "No escribiste ningún valor nuevo, no se"
                                " guardó nada."
                            )
                        else:
                            try:
                                guardar_configuracion_sistema(
                                    supabase,
                                    st.session_state.empresa_id,
                                    **campos_a_guardar,
                                )
                                for campo, valor in (
                                    campos_a_guardar.items()
                                ):
                                    setattr(
                                        st.session_state, campo, valor
                                    )
                                st.success(
                                    "✅ Configuración de seguridad"
                                    " actualizada y guardada en Supabase."
                                )
                            except Exception as e_cfg:
                                st.error(
                                    f"No se pudo guardar en Supabase:"
                                    f" {e_cfg}"
                                )